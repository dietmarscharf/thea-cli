#!/usr/bin/env python3
"""
Depotkonto.py - Analyse und Markdown-Generierung für Depotkonten
Verarbeitet BLUEITS-Depotkonto-7274079 und Ramteid-Depotkonto-7274087
"""

import os
import re
from pathlib import Path
from typing import Dict, List, Any
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

from Konten import BaseKontoAnalyzer, calculate_monthly_aggregates


class DepotkontoAnalyzer(BaseKontoAnalyzer):
    def __init__(self):
        super().__init__()
        self.depots = {
            "BLUEITS": {
                "folder": "BLUEITS-Depotkonto-7274079",
                "depot_number": "7274079",
                "output_file": "BLUEITS-Depotkonto.html",
                "company_name": "BLUEITS GmbH",
                "fiscal_year": {
                    "type": "april_march",
                    "description": "April bis März",
                    "example": "Geschäftsjahr 2023 = 01.04.2023 - 31.03.2024"
                }
            },
            "Ramteid": {
                "folder": "Ramteid-Depotkonto-7274087",
                "depot_number": "7274087",
                "output_file": "Ramteid-Depotkonto.html",
                "company_name": "Ramteid GmbH",
                "fiscal_year": {
                    "type": "calendar",
                    "description": "Kalenderjahr",
                    "example": "Geschäftsjahr 2023 = 01.01.2023 - 31.12.2023"
                }
            }
        }
    
    def is_stock_isin(self, isin: str) -> bool:
        """
        Bestimmt ob eine ISIN eine Aktie oder ein Nicht-Aktien-Produkt ist.
        
        Aktien:
        - US88160R1014 (TESLA INC)
        - Weitere US-Aktien (US...)
        
        Nicht-Aktien (Derivate, Zertifikate, strukturierte Produkte):
        - DE000JN9UFS3 (J.P. Morgan Structured Product)
        - Andere strukturierte Produkte und Derivate
        """
        if not isin:
            return False
            
        # Tesla und andere US-Aktien sind Aktien
        if isin.startswith('US'):
            return True
            
        # DE-ISINs können Aktien oder Derivate sein
        # Bekannte Derivate/Zertifikate explizit als Nicht-Aktien markieren
        known_derivatives = [
            'DE000JN9UFS3',  # J.P. Morgan Structured Product
            # Weitere bekannte Derivate können hier hinzugefügt werden
        ]
        
        if isin in known_derivatives:
            return False
            
        # Standard: DE-ISINs als Aktien behandeln, außer sie sind als Derivate bekannt
        # Dies kann je nach Portfolio angepasst werden
        if isin.startswith('DE'):
            # Wenn nicht in der Liste der bekannten Derivate, als Aktie behandeln
            return isin not in known_derivatives
            
        # Andere Länder-ISINs (FR, GB, etc.) sind üblicherweise Aktien
        return True
    
    def get_fiscal_year(self, date_str: str, fiscal_type: str) -> str:
        """
        Bestimmt das Geschäftsjahr basierend auf Datum und Geschäftsjahr-Typ.
        
        Args:
            date_str: Datum im Format YYYY-MM-DD oder als String
            fiscal_type: 'april_march' oder 'calendar'
            
        Returns:
            Geschäftsjahr als String, z.B. 'FY2024'
        """
        import datetime
        
        if not date_str:
            return 'FY0000'
        
        # Parse date if it's a string
        if isinstance(date_str, str):
            try:
                # Try parsing YYYY-MM-DD format
                if '-' in date_str:
                    date_obj = datetime.datetime.strptime(date_str, '%Y-%m-%d')
                # Try parsing DD.MM.YYYY format
                elif '.' in date_str:
                    date_obj = datetime.datetime.strptime(date_str, '%d.%m.%Y')
                else:
                    return 'FY0000'
            except:
                return 'FY0000'
        else:
            date_obj = date_str
        
        year = date_obj.year
        month = date_obj.month
        
        if fiscal_type == 'april_march':
            # April-März: Wenn Monat >= 4 (April), dann gehört es zum GJ des nächsten Jahres
            # z.B. 15.05.2024 -> FY2025 (GJ April 2024 - März 2025)
            if month >= 4:
                return f'FY{year + 1}'
            else:
                return f'FY{year}'
        else:  # calendar year
            return f'FY{year}'
    
    def get_calendar_year(self, date_str: str) -> str:
        """
        Bestimmt das Kalenderjahr aus einem Datum.
        
        Args:
            date_str: Datum im Format YYYY-MM-DD oder als String
            
        Returns:
            Kalenderjahr als String, z.B. 'CY2024'
        """
        import datetime
        
        if not date_str:
            return 'CY0000'
        
        # Parse date if it's a string
        if isinstance(date_str, str):
            try:
                # Try parsing YYYY-MM-DD format
                if '-' in date_str:
                    date_obj = datetime.datetime.strptime(date_str, '%Y-%m-%d')
                # Try parsing DD.MM.YYYY format
                elif '.' in date_str:
                    date_obj = datetime.datetime.strptime(date_str, '%d.%m.%Y')
                else:
                    return 'CY0000'
            except:
                return 'CY0000'
        else:
            date_obj = date_str
        
        return f'CY{date_obj.year}'
    
    def extract_trading_details(self, text: str) -> Dict[str, Any]:
        """
        Extrahiert detaillierte Handelsinformationen aus Orderabrechnungen.
        
        Wichtige Hinweise zu Gebühren:
        - Viele Transaktionen haben legitim KEINE Gebühren (wenn Kurswert = Ausmachender Betrag)
        - Dies tritt häufig auf bei: Sparplänen, Aktionsangeboten, Teilausführungen
        - Typische Gebührenstruktur wenn vorhanden:
          * Provision: 0,25% - 1% des Ordervolumens
          * Handelsplatzgebühr: 2-5 EUR
          * Mindestgebühr: 8-10 EUR  
          * USt: 19% auf alle Gebühren (bereits in Gebührenbetrag enthalten)
        """
        details = {
            'shares': None,
            'execution_price': None,
            'execution_price_min': None,
            'execution_price_max': None,
            'execution_price_avg': None,
            'limit_price': None,
            'gross_amount': None,
            'fees': None,
            'is_vat_free': False,    # Flag für umsatzsteuerfreie Gebühren
            'profit_loss': None,
            'net_amount': None,
            'execution_date': None,  # Schlusstag/Handelstag (Ausführungsdatum)
            'value_date': None,      # Valuta/Wertstellung (Settlement)
            'booking_date': None     # Buchungsdatum (falls abweichend)
        }
        
        # Extrahiere Stückzahl (mit Unterstützung für deutsche Tausendertrennzeichen)
        shares_match = re.search(r'Stück\s+([\d.]+)', text)
        if shares_match:
            # Entferne deutsche Tausendertrennzeichen (Punkt) vor der Konvertierung
            shares_str = shares_match.group(1).replace('.', '')
            details['shares'] = int(shares_str)
        
        # Extrahiere Limit (falls vorhanden)
        limit_match = re.search(r'Limit\s+([\d.,]+)\s*EUR', text)
        if limit_match:
            details['limit_price'] = float(limit_match.group(1).replace('.', '').replace(',', '.'))
        
        # Extrahiere Ausführungskurs(e) und dedupliziere
        exec_prices = re.findall(r'Ausführungskurs\s+([\d.,]+)\s*EUR', text)
        if exec_prices:
            # Konvertiere und dedupliziere Kurse
            prices = list(set([float(p.replace('.', '').replace(',', '.')) for p in exec_prices]))
            prices.sort()  # Sortiere für konsistente Min/Max
            
            if len(prices) == 1:
                details['execution_price'] = prices[0]
            else:
                # Nur bei tatsächlich verschiedenen Kursen
                details['execution_price_min'] = prices[0]
                details['execution_price_max'] = prices[-1]
                details['execution_price_avg'] = sum(prices) / len(prices)
        
        # Extrahiere Kurswert (Brutto)
        # Erweiterte Patterns für verschiedene Layouts
        kurswert_patterns = [
            r'Kurswert\s+([\d.,]+)[\s-]*EUR',  # Standard mit EUR
            r'Kurswert\s+([\d.,]+)\s*\n\s*EUR',  # EUR auf nächster Zeile
            r'Kurswert[\s\n]+([\d.,]+)',  # Ohne EUR-Anforderung
            r'Kurswert\s*\|\s*([\d.,]+)',  # Mit Pipe-Separator
        ]
        
        for pattern in kurswert_patterns:
            kurswert_match = re.search(pattern, text, re.MULTILINE | re.DOTALL)
            if kurswert_match:
                value_str = kurswert_match.group(1).replace('.', '').replace(',', '.')
                try:
                    value = float(value_str)
                    if value > 0:  # Plausibilitätsprüfung
                        details['gross_amount'] = value
                        break
                except ValueError:
                    continue
        
        # Extrahiere Provision/Gebühren mit erweiterten Patterns
        fee_patterns = [
            # Spezifische Gebührenarten
            r'Provision\s+([\d.,]+)[\s-]*EUR',
            r'Handelsplatzgebühr\s+([\d.,]+)[\s-]*EUR',
            r'Maklercourtage\s+([\d.,]+)[\s-]*EUR',
            r'Fremdspesen\s+([\d.,]+)[\s-]*EUR',
            r'Börsengebühr\s+([\d.,]+)[\s-]*EUR',
            r'Transaktionsentgelt\s+([\d.,]+)[\s-]*EUR',
            r'Orderprovision\s+([\d.,]+)[\s-]*EUR',
            # Summen und Gesamtbeträge
            r'Gebühren\s+gesamt\s+([\d.,]+)[\s-]*EUR',
            r'Summe\s+Gebühren\s+([\d.,]+)[\s-]*EUR',
            r'Provision\s+und\s+Gebühren\s+([\d.,]+)[\s-]*EUR',
            # Mit Pipe-Separator
            r'Provision\s*\|\s*([\d.,]+)\s*€',
            r'Gebühren\s*\|\s*([\d.,]+)\s*€',
            r'Handelsplatzgebühr\s*\|\s*([\d.,]+)\s*€',
            r'Summe\s*\|\s*([\d.,]+)\s*€',
            # Minus-Beträge (Gebühren als Abzug)
            r'[\-\s]+([\d.,]+)\s*EUR\s*(?:Provision|Gebühr)',
            r'Abzüge\s+([\d.,]+)\s*EUR',
            # Allgemeinere Patterns
            r'Gebühren[^\d]+([\d.,]+)\s*(?:EUR|€)',
            r'Entgelte[^\d]+([\d.,]+)\s*(?:EUR|€)',
            r'Kosten[^\d]+([\d.,]+)\s*(?:EUR|€)'
        ]
        
        total_fees = 0.0
        fees_found = []
        
        # Suche nach allen Gebührenposten und summiere sie
        for pattern in fee_patterns:
            matches = re.findall(pattern, text)
            for match in matches:
                fee_value = float(match.replace('.', '').replace(',', '.'))
                if fee_value > 0 and fee_value < 10000:  # Plausibilitätsprüfung
                    fees_found.append(fee_value)
        
        # Verwende die höchste gefundene Gebühr (wahrscheinlich die Summe)
        if fees_found:
            details['fees'] = max(fees_found)
        
        # Check if fees are VAT-free
        vat_free_patterns = [
            r'umsatzsteuerbefreite\s+Finanzdienstleistung',
            r'umsatzsteuerfrei',
            r'ohne\s+Umsatzsteuer',
            r'Sofern\s+keine\s+Umsatzsteuer\s+ausgewiesen',
            r'Börsenplatzentgelt',  # Always VAT-free
            r'Maklercourtage'       # Always VAT-free
        ]
        
        for pattern in vat_free_patterns:
            if re.search(pattern, text, re.IGNORECASE):
                details['is_vat_free'] = True
                break
        
        # Extrahiere Veräußerungsgewinn/Verlust mit erweiterten Patterns
        # Spezialbehandlung für vertikales Layout (Sparkasse-Format)
        vertical_layout_match = re.search(
            r'Ermittlung steuerrelevante Erträge\s+Veräußerungs(?:gewinn|verlust)\s+Ausmachender Betrag',
            text, re.IGNORECASE
        )
        
        if vertical_layout_match:
            # Bei vertikalem Layout: Erste Zahl = Gewinn/Verlust, Zweite Zahl = Ausmachender Betrag
            lines = text[vertical_layout_match.end():].split('\n')
            numbers_found = []
            for i, line in enumerate(lines[:15]):  # Schaue die nächsten 15 Zeilen
                # Suche nach Zahlen mit oder ohne EUR in derselben oder nächsten Zeilen
                # Unterstützt auch nachgestelltes Minus (deutsches Buchführungsformat)
                number_match = re.search(r'([-]?[\d]{1,3}(?:\.[\d]{3})*(?:,[\d]{2})?[-]?)', line)
                if number_match:
                    # Prüfe ob EUR in derselben oder in den nächsten 2 Zeilen ist
                    # (EUR kann auf einer separaten Zeile stehen)
                    has_eur = ('EUR' in line or 
                              (i+1 < len(lines) and 'EUR' in lines[i+1]) or
                              (i+2 < len(lines) and 'EUR' in lines[i+2]))
                    if has_eur:
                        # Prüfe ob es eine plausible Zahl ist
                        # Behandle nachgestelltes Minus
                        raw_number = number_match.group(1)
                        test_str = raw_number.rstrip('-').replace('.', '').replace(',', '.')
                        try:
                            test_val = float(test_str)
                            if test_val > 100:  # Ignoriere kleine Zahlen
                                numbers_found.append(raw_number)
                        except:
                            pass
                if len(numbers_found) >= 2:
                    break
            
            if len(numbers_found) >= 2:
                # Konvertiere beide Zahlen mit Unterstützung für nachgestelltes Minus
                num1_str = numbers_found[0]
                num2_str = numbers_found[1]
                
                # Prüfe und behandle nachgestelltes Minus
                value1_sign = -1.0 if num1_str.endswith('-') else 1.0
                value2_sign = -1.0 if num2_str.endswith('-') else 1.0
                
                value1 = float(num1_str.rstrip('-').replace('.', '').replace(',', '.')) * value1_sign
                value2 = float(num2_str.rstrip('-').replace('.', '').replace(',', '.')) * value2_sign
                
                # Debug-Ausgabe für BLUEITS-Transaktionen
                if 'BLUEITS' in text[:1000] and 'Verkauf' in text[:1000]:
                    print(f"  Debug: Vertikales Layout - Wert1={value1}, Wert2={value2}")
                    if details.get('gross_amount'):
                        print(f"  Debug: Kurswert={details['gross_amount']}")
                
                # Verbesserte Zuordnungslogik
                # Prinzip: Der Ausmachende Betrag ist IMMER ähnlich zum Kurswert
                # Der Gewinn/Verlust ist IMMER deutlich kleiner
                
                if details.get('gross_amount'):
                    kurswert = details['gross_amount']
                    
                    # Berechne absolute Differenzen zum Kurswert
                    diff1_to_kurswert = abs(value1 - kurswert)
                    diff2_to_kurswert = abs(value2 - kurswert)
                    
                    # Der Wert mit der kleineren Differenz zum Kurswert ist der Ausmachende Betrag
                    if diff1_to_kurswert < diff2_to_kurswert:
                        # value1 ist näher am Kurswert -> Ausmachender Betrag
                        details['net_amount'] = value1
                        details['profit_loss'] = value2
                        if 'BLUEITS' in text[:1000]:
                            print(f"  -> Zuordnung: Ausmachend={value1:.2f}, G/V={value2:.2f}")
                    else:
                        # value2 ist näher am Kurswert -> Ausmachender Betrag  
                        details['net_amount'] = value2
                        details['profit_loss'] = value1
                        if 'BLUEITS' in text[:1000]:
                            print(f"  -> Zuordnung: Ausmachend={value2:.2f}, G/V={value1:.2f}")
                else:
                    # Ohne Kurswert: Der größere Wert ist fast immer der Ausmachende Betrag
                    # (da Gewinn/Verlust normalerweise nur ein Bruchteil des Transaktionswerts ist)
                    if value2 > value1 * 3:  # Wenn value2 deutlich größer
                        details['profit_loss'] = value1
                        details['net_amount'] = value2
                    elif value1 > value2 * 3:  # Wenn value1 deutlich größer
                        details['profit_loss'] = value2
                        details['net_amount'] = value1
                    else:
                        # Bei ähnlichen Werten: Dokumentreihenfolge (erst G/V, dann Ausmachend)
                        details['profit_loss'] = value1
                        details['net_amount'] = value2
        else:
            # Standard Patterns für normale Layouts
            profit_loss_patterns = [
                # Patterns mit nachgestelltem Minus (deutsches Buchführungsformat)
                (r'Veräußerungsverlust\s+([\d.,]+)-\s*EUR', -1.0),  # Verlust mit nachgestelltem Minus
                (r'Veräußerungsgewinn\s+([\d.,]+)\s*EUR', 1.0),     # Gewinn ohne Minus
                # Standard Patterns (mit vorangestelltem Minus)
                (r'Veräußerungsgewinn\s+([-]?[\d.,]+)\s*EUR', 1.0),  # Positiv
                (r'Veräußerungsverlust\s+([-]?[\d.,]+)\s*EUR', -1.0), # Negativ
                # Alternative Begriffe
                (r'Gewinn\s+aus\s+Verkauf\s+([\d.,]+)-?\s*EUR', lambda x: -1.0 if x.endswith('-') else 1.0),
                (r'Verlust\s+aus\s+Verkauf\s+([\d.,]+)-?\s*EUR', -1.0),
                (r'Realisierter\s+Gewinn\s+([\d.,]+)-?\s*EUR', lambda x: -1.0 if x.endswith('-') else 1.0),
                (r'Realisierter\s+Verlust\s+([\d.,]+)-?\s*EUR', -1.0),
                # Mit Pipe-Separator
                (r'Veräußerungsgewinn\s*\|\s*([-]?[\d.,]+)\s*€', 1.0),
                (r'Veräußerungsverlust\s*\|\s*([-]?[\d.,]+)\s*€', -1.0),
                # Spezielle Formate (Betrag in Klammern = negativ)
                (r'Veräußerungsergebnis\s+\(([\d.,]+)\)\s*EUR', -1.0),  # In Klammern = Verlust
                (r'Veräußerungsergebnis\s+([\d.,]+)\s*EUR', 1.0),       # Ohne Klammern = Gewinn
                # Allgemeinere Patterns
                (r'Gewinn/Verlust\s+([-]?[\d.,]+)\s*(?:EUR|€)', 1.0),
                (r'Ergebnis\s+([-]?[\d.,]+)\s*(?:EUR|€)', 1.0)
            ]
            
            for pattern, multiplier in profit_loss_patterns:
                match = re.search(pattern, text)
                if match:
                    # Extrahiere den Wert und entferne nachgestelltes Minus
                    value_str = match.group(1).rstrip('-')
                    # Prüfe ob ursprünglich ein nachgestelltes Minus vorhanden war
                    has_trailing_minus = match.group(1).endswith('-')
                    # Konvertiere zu float (deutsches Format)
                    value = float(value_str.replace('.', '').replace(',', '.'))
                    # Wende Multiplier an, berücksichtige nachgestelltes Minus
                    if callable(multiplier):
                        details['profit_loss'] = value * (multiplier(match.group(1)))
                    else:
                        # Bei nachgestelltem Minus ist der Wert immer negativ
                        if has_trailing_minus and multiplier > 0:
                            details['profit_loss'] = -value
                        else:
                            details['profit_loss'] = value * multiplier
                    break
        
        # Extrahiere Ausmachender Betrag (Netto) - nur wenn nicht bereits aus vertikalem Layout extrahiert
        if not details.get('net_amount'):
            ausmachend_patterns = [
                r'Ausmachender\s+Betrag\s+([\d.,]+)\s*EUR',
                r'Zu\s+Ihren\s+Gunsten\s+([\d.,]+)\s*EUR',
                r'Zu\s+Ihren\s+Lasten\s+([\d.,]+)\s*EUR',
                r'Abrechnungsbetrag\s+([\d.,]+)\s*EUR',
                r'Endbetrag\s+([\d.,]+)\s*EUR',
                r'Gesamt\s+([\d.,]+)\s*EUR',
                # Mit Pipe-Separator
                r'Ausmachender\s+Betrag\s*\|\s*([\d.,]+)\s*€',
                r'Gesamt\s*\|\s*([\d.,]+)\s*€'
            ]
            
            for pattern in ausmachend_patterns:
                match = re.search(pattern, text)
                if match:
                    details['net_amount'] = float(match.group(1).replace('.', '').replace(',', '.'))
                    break
        
        # Extrahiere Vormerkungsentgelt (für Limit-Order-Vormerkungen)
        vormerkung_pattern = r'Vormerkungsentgelt\s+von\s+([\d,]+)\s*EUR'
        vormerkung_match = re.search(vormerkung_pattern, text)
        if vormerkung_match:
            details['fees'] = float(vormerkung_match.group(1).replace(',', '.'))
            details['is_vat_free'] = True  # Vormerkungsentgelt ist ohne USt
        
        # Fallback: Berechne Gebühren aus Differenz wenn möglich
        if details['gross_amount'] and details['net_amount'] and not details['fees']:
            # Bei Verkauf: Gebühren = Kurswert - Ausmachend
            # Bei Kauf: Gebühren = Ausmachend - Kurswert
            diff = abs(details['gross_amount'] - details['net_amount'])
            if diff > 0 and diff < details['gross_amount'] * 0.1:  # Max 10% Gebühren
                details['fees'] = diff
        
        # Validierung der extrahierten Werte
        if details.get('gross_amount') and details.get('net_amount'):
            # Bei Verkäufen sollte net_amount >= gross_amount - fees sein
            if details.get('profit_loss') is not None:
                # Prüfe ob der Gewinn/Verlust plausibel ist
                if abs(details['profit_loss']) > details['gross_amount']:
                    # Gewinn/Verlust kann nicht größer als der Kurswert sein
                    # Möglicherweise wurden die Werte vertauscht
                    print(f"  ⚠️  Warnung: Unplausibler G/V-Wert: {details['profit_loss']} bei Kurswert {details['gross_amount']}")
            
            # Prüfe ob Ausmachender Betrag plausibel ist
            if details['net_amount'] < details['gross_amount'] * 0.5:
                # Ausmachender Betrag sollte nicht weniger als 50% des Kurswerts sein (außer bei hohen Gebühren)
                if not details.get('fees') or details['fees'] < details['gross_amount'] * 0.5:
                    print(f"  ⚠️  Warnung: Unplausibler Ausmachender Betrag: {details['net_amount']} bei Kurswert {details['gross_amount']}")
        
        # Extrahiere Schlusstag/Handelstag (Ausführungsdatum)
        execution_patterns = [
            r'Schlusstag/-Zeit\s+(\d{1,2})\.(\d{1,2})\.(\d{4})',  # Mit Zeit
            r'Schlusstag\s*[:\s]\s*(\d{1,2})\.(\d{1,2})\.(\d{4})',
            r'Schlusstag\s+(\d{1,2})\.(\d{1,2})\.(\d{4})',
            r'Ausführungstag\s*[:\s]\s*(\d{1,2})\.(\d{1,2})\.(\d{4})',
            r'Ausführung\s+am\s+(\d{1,2})\.(\d{1,2})\.(\d{4})',
            r'Geschäftstag\s*[:\s]\s*(\d{1,2})\.(\d{1,2})\.(\d{4})',
            r'Handelstag\s*[:\s]\s*(\d{1,2})\.(\d{1,2})\.(\d{4})',
            # Mit Pipe-Separator
            r'Schlusstag\s*\|\s*(\d{1,2})\.(\d{1,2})\.(\d{4})',
            r'Geschäftstag\s*\|\s*(\d{1,2})\.(\d{1,2})\.(\d{4})',
            # In Tabellen
            r'Ausführ\.-tag[^|]*\|\s*(\d{1,2})\.(\d{1,2})\.(\d{4})',
        ]
        
        for pattern in execution_patterns:
            match = re.search(pattern, text)
            if match:
                day, month, year = match.groups()
                details['execution_date'] = f"{year}-{month.zfill(2)}-{day.zfill(2)}"
                break
        
        # Extrahiere Valuta/Wertstellung (Settlement-Datum)
        valuta_patterns = [
            r'mit\s+Valuta\s+(\d{1,2})\.(\d{1,2})\.(\d{4})',  # "mit Valuta DD.MM.YYYY"
            r'Valuta\s*[:\s]\s*(\d{1,2})\.(\d{1,2})\.(\d{4})',
            r'Valuta\s+(\d{1,2})\.(\d{1,2})\.(\d{4})',
            r'Wertstellung\s*[:\s]\s*(\d{1,2})\.(\d{1,2})\.(\d{4})',
            # Mit Pipe-Separator
            r'Valuta\s*\|\s*(\d{1,2})\.(\d{1,2})\.(\d{4})',
            r'Wertstellung\s*\|\s*(\d{1,2})\.(\d{1,2})\.(\d{4})',
        ]
        
        for pattern in valuta_patterns:
            match = re.search(pattern, text)
            if match:
                day, month, year = match.groups()
                details['value_date'] = f"{year}-{month.zfill(2)}-{day.zfill(2)}"
                break
        
        # Extrahiere Buchungsdatum (falls explizit angegeben)
        booking_patterns = [
            r'Buchungsdatum\s*[:\s]\s*(\d{1,2})\.(\d{1,2})\.(\d{4})',
            r'Buchung\s+am\s+(\d{1,2})\.(\d{1,2})\.(\d{4})',
            r'gebucht\s+am\s+(\d{1,2})\.(\d{1,2})\.(\d{4})',
        ]
        
        for pattern in booking_patterns:
            match = re.search(pattern, text)
            if match:
                day, month, year = match.groups()
                details['booking_date'] = f"{year}-{month.zfill(2)}-{day.zfill(2)}"
                break
        
        # Fallback für Stichtag (für Kompatibilität und Depotabschlüsse)
        if not details['execution_date']:
            stichtag_patterns = [
                r'Stichtag\s*[:\s]\s*(\d{1,2})\.(\d{1,2})\.(\d{4})',
                r'per\s+(\d{1,2})\.(\d{1,2})\.(\d{4})',
                r'Stand\s*[:\s]\s*(\d{1,2})\.(\d{1,2})\.(\d{4})',
            ]
            
            for pattern in stichtag_patterns:
                match = re.search(pattern, text)
                if match:
                    day, month, year = match.groups()
                    details['stichtag'] = f"{year}-{month.zfill(2)}-{day.zfill(2)}"
                    break
        else:
            # Für Rückwärtskompatibilität: setze stichtag = execution_date
            details['stichtag'] = details['execution_date']
        
        # Extrahiere Depotentgelte (für Depotabschlüsse)
        depot_fee_patterns = [
            # Standard Format: "netto X,XX Euro + Y% USt Z,ZZ Euro = brutto A,AA Euro"
            (r'Depotentgelte.*?netto\s+([\d,]+)\s*Euro.*?(\d+)%\s*USt\s+([\d,]+)\s*Euro.*?brutto\s+([\d,]+)\s*Euro', 'full'),
            # Alternatives Format ohne "Depotentgelte" prefix
            (r'netto\s+([\d,]+)\s*Euro\s*\+\s*(\d+)%\s*USt\s+([\d,]+)\s*Euro\s*=\s*brutto\s+([\d,]+)\s*Euro', 'full'),
            # Format mit "Die Depotentgelte ... betragen" (fixed: use .*? instead of [^b]*?)
            (r'Die\s+Depotentgelte.*?betragen\s+netto\s+([\d,]+)\s+Euro\s+\+\s+(\d+)%\s+USt\s+([\d,]+)\s+Euro\s+=\s+brutto\s+([\d,]+)\s+Euro', 'full'),
            # Pattern mit mehr Flexibilität für Whitespace
            (r'betragen\s+netto\s+([\d,]+)\s*Euro\s*\+\s*(\d+)%\s*USt\s+([\d,]+)\s*Euro\s*=\s*brutto\s+([\d,]+)\s*Euro', 'full'),
            # Nur Bruttobetrag
            (r'Depotentgelte.*?betragen.*?([\d,]+)\s*Euro', 'gross_only'),
            # Mit "Die Depotentgelte"
            (r'Die\s+Depotentgelte.*?betragen.*?brutto\s+([\d,]+)\s*Euro', 'gross_only'),
        ]
        
        for pattern, pattern_type in depot_fee_patterns:
            match = re.search(pattern, text, re.DOTALL | re.IGNORECASE)
            if match:
                if pattern_type == 'full' and len(match.groups()) == 4:
                    # Vollständige Information mit Netto, USt%, USt-Betrag und Brutto
                    details['depot_fee_net'] = float(match.group(1).replace(',', '.'))
                    details['vat_rate'] = int(match.group(2))
                    details['vat_amount'] = float(match.group(3).replace(',', '.'))
                    details['fees'] = float(match.group(4).replace(',', '.'))  # Brutto
                    details['is_depot_fee'] = True
                    details['is_vat_free'] = False  # Depotentgelte haben USt
                elif pattern_type == 'gross_only':
                    # Nur Bruttobetrag
                    details['fees'] = float(match.group(1).replace(',', '.'))
                    details['is_depot_fee'] = True
                    details['is_vat_free'] = False  # Depotentgelte haben normalerweise USt
                break
        
        return details
    
    def extract_transaction_data(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """Extrahiert depot-spezifische Transaktionsdaten"""
        base_data = self.get_base_transaction_data(data)
        if not base_data:
            return None
            
        extracted_text = base_data['extracted_text']
        
        # Check if THEA extraction failed
        if not extracted_text or (data.get('response', {}).get('json') is None and data.get('errors')):
            # THEA extraction failed - create warning entry
            result = {
                'extraction_failed': True,
                'type': '⚠️ EXTRAKTION-FEHLGESCHLAGEN',
                'doc_date': base_data.get('doc_date'),
                'file_date': base_data.get('file_date'),
                'file': base_data.get('file'),
                'pdf_path': base_data.get('pdf_path'),
                'original_file': base_data.get('original_file'),  # Include original filename for HTML link
                'date': base_data.get('doc_date'),
                'isin': None,
                'shares': None,
                'purchase_value': None,
                'fees': None,
                'total_amount': None,
                'profit_loss': None,
                'errors': data.get('errors', [])
            }
            return result
        
        # Extrahiere ISIN
        isin_match = re.search(r'([A-Z]{2}[A-Z0-9]{10})', extracted_text)
        isin = isin_match.group(1) if isin_match else None
        
        # Extrahiere detaillierte Handelsdaten
        trading_details = self.extract_trading_details(extracted_text)
        
        # Fallback auf alte Methode für Kompatibilität
        amounts = self.extract_amounts_from_text(extracted_text)
        
        # Bestimme Transaktionstyp und Periode
        transaction_type = 'Unbekannt'
        period = None
        extracted_lower = extracted_text.lower()
        filename_lower = base_data.get('original_file', '').lower()
        date_str = base_data.get('date', '')
        
        # Check for execution notices (preliminary confirmations)
        if 'ausführungsanzeige' in extracted_lower:
            if 'verkauf' in extracted_lower:
                transaction_type = 'Ausführungsanzeige-Verkauf'
            elif 'kauf' in extracted_lower:
                transaction_type = 'Ausführungsanzeige-Kauf'
            else:
                transaction_type = 'Ausführungsanzeige'
        # PRIORITY CHECK: New format from April 2022 onwards - "Wertpapier Abrechnung Verkauf/Kauf"
        # This must be checked early to ensure correct classification
        elif 'wertpapier abrechnung verkauf' in extracted_lower:
            transaction_type = 'Orderabrechnung-Verkauf'
        elif 'wertpapier abrechnung kauf' in extracted_lower:
            transaction_type = 'Orderabrechnung-Kauf'
        # Check for capital measures (stock splits, etc.)
        elif ('kapitalmaßnahme' in extracted_lower or 'kapitalmassnahme' in extracted_lower or
            'aktiensplit' in extracted_lower or 'stock split' in extracted_lower or
            'neueinteilung des grundkapitals' in extracted_lower or
            'kapitalmaßnahme' in filename_lower or 'kapitalmassnahme' in filename_lower):
            transaction_type = 'Kapitalmaßnahme'
        # Check for cost information documents (MiFID II Kostenaufstellung)
        # These are specifically the Ex-Post cost reports, not regular orders mentioning cost information
        elif ('information über kosten und nebenkosten' in extracted_lower or 
            'ex-post-kosteninformation' in extracted_lower or
            ('dienstleistungskosten' in extracted_lower and 'übergreifende kosten' in extracted_lower)):
            transaction_type = 'Kostenaufstellung'
        # Check for depot statements (various formats)
        elif any(term in extracted_lower for term in ['depotabschluss', 'jahresabschluss', 'jahresdepotauszug', 'depotauszug', 'ex-post-rep']):
            # Determine if annual or quarterly
            if 'jahresabschluss' in extracted_lower or 'jahres' in extracted_lower:
                transaction_type = 'Depotabschluss-Jahresabschluss'
                period = 'Jährlich'
            elif date_str and len(date_str) >= 7:
                month = int(date_str[5:7]) if date_str[5:7].isdigit() else 0
                if month == 1:
                    transaction_type = 'Depotabschluss-Jahresabschluss'
                    period = 'Jährlich'
                elif month in [4, 5]:
                    transaction_type = 'Depotabschluss-Quartal'
                    period = 'Q1'
                elif month == 7:
                    transaction_type = 'Depotabschluss-Quartal'
                    period = 'Q2'
                elif month == 10:
                    transaction_type = 'Depotabschluss-Quartal'
                    period = 'Q3'
                else:
                    transaction_type = 'Depotabschluss'
            else:
                transaction_type = 'Depotabschluss'
        # Fallback: check filename for depot statement indicators
        elif any(term in filename_lower for term in ['depotabschluss', 'ex-post-rep', 'jahresabschluss']):
            if 'jahresabschluss' in filename_lower:
                transaction_type = 'Depotabschluss-Jahresabschluss'
                period = 'Jährlich'
            else:
                transaction_type = 'Depotabschluss'
        # Check for Orderabrechnung/Wertpapierabrechnung
        elif 'orderabrechnung' in filename_lower or 'wertpapier' in extracted_lower:
            # Check for limit order confirmation (Vormerkung) first
            if 'auftragsbestätigung' in extracted_lower or 'vormerkungsentgelt' in extracted_lower:
                transaction_type = 'Limit-Order-Vormerkung'
            # Check for order cancellation
            elif 'streichungsbestätigung' in extracted_lower or 'order wurde' in extracted_lower and 'gestrichen' in extracted_lower:
                transaction_type = 'Orderabrechnung-Stornierung'
            # Determine if buy or sell
            elif 'verkauf' in extracted_lower or 'sell' in extracted_lower:
                transaction_type = 'Orderabrechnung-Verkauf'
            elif 'kauf' in extracted_lower or 'buy' in extracted_lower or 'erwerb' in extracted_lower:
                transaction_type = 'Orderabrechnung-Kauf'
            else:
                # Try to detect from "Wertpapier Abrechnung" header
                if 'wertpapier abrechnung verkauf' in extracted_lower:
                    transaction_type = 'Orderabrechnung-Verkauf'
                elif 'wertpapier abrechnung kauf' in extracted_lower:
                    transaction_type = 'Orderabrechnung-Kauf'
                else:
                    transaction_type = 'Orderabrechnung'
        # Standard buy/sell detection  
        elif 'verkauf' in extracted_lower or 'sell' in extracted_lower:
            transaction_type = 'Verkauf'
        elif 'kauf' in extracted_lower or 'buy' in extracted_lower or 'erwerb' in extracted_lower:
            transaction_type = 'Kauf'
            
        # Special handling for failed extractions
        if transaction_type == 'Unbekannt':
            # Check if it's a failed Orderabrechnung based on filename
            if 'orderabrechnung' in filename_lower:
                transaction_type = 'Orderabrechnung-Fehlgeschlagen'
            
        # Extrahiere Transaktionsreferenznummer (für korrekte Sortierung bei mehreren Transaktionen am selben Tag)
        transaction_ref = None
        order_number = None
        
        # Suche nach Rechnungsnummer-Pattern (z.B. W02279-0000013797/22)
        ref_match = re.search(r'(W\d+-\d+/\d+)', extracted_text)
        if ref_match:
            transaction_ref = ref_match.group(1)
        
        # Suche nach Auftragsnummer-Pattern (z.B. 768384/63.00)
        order_match = re.search(r'(\d{6}/\d+\.\d+)', extracted_text)
        if order_match:
            order_number = order_match.group(1)
        
        # Bestimme Wertpapiertyp (Aktie oder Nicht-Aktie)
        security_type = 'stock' if self.is_stock_isin(isin) else 'non-stock' if isin else None
        
        # Handle stock splits (Kapitalmaßnahme)
        stock_split_data = {}
        if transaction_type == 'Kapitalmaßnahme':
            # Extract stock split data
            split_data = self.extract_stock_split_data(extracted_text)
            if split_data:
                stock_split_data['original_shares'] = split_data.get('original_shares', 0)
                stock_split_data['split_multiplier'] = split_data.get('split_multiplier', 1)
                stock_split_data['new_shares_added'] = split_data.get('new_shares_added', 0)
                stock_split_data['split_ratio_text'] = split_data.get('split_ratio_text', '')
                # Set shares to original amount (before split)
                if 'original_shares' in split_data:
                    trading_details['shares'] = split_data['original_shares']
        
        # Kombiniere alle Daten
        result = {
            **base_data,
            'type': transaction_type,
            'period': period,
            'isin': isin,
            'security_type': security_type,  # Neu: Typ des Wertpapiers
            'transaction_ref': transaction_ref,  # Neu: Transaktionsreferenz für Sortierung
            'order_number': order_number,  # Neu: Auftragsnummer
            'amounts': amounts,
            'max_amount': max(amounts) if amounts else 0,
            # Neue detaillierte Handelsdaten
            **trading_details,
            # Stock split data (if applicable)
            **stock_split_data
        }
        
        # Verwende net_amount als max_amount wenn verfügbar (genauer)
        if trading_details.get('net_amount'):
            result['max_amount'] = trading_details['net_amount']
        elif trading_details.get('gross_amount'):
            result['max_amount'] = trading_details['gross_amount']
        
        # DEPRECATED: Fallback für Depotabschluss-Gebühren 
        # Gebühren werden jetzt über depot_statement_lookup übertragen (siehe Zeilen 1408-1414 und 2311-2317)
        # Dieser Code wird nicht mehr benötigt, da extract_depot_balance() die Gebühren extrahiert
        # und sie über depot_statement_lookup an die Transaktionen weitergibt.
        #
        # if 'Depotabschluss' in transaction_type and not result.get('fees'):
        #     import glob
        #     import os
        #     original_file = base_data.get('original_file', '')
        #     if original_file:
        #         # Suche nach docling.txt Datei
        #         base_name = original_file.replace('.pdf', '.pdf')
        #         docling_pattern = f"{base_name}.*docling.txt"
        #         docling_files = glob.glob(docling_pattern)
        #         
        #         if docling_files:
        #             # Verwende die neueste docling.txt Datei
        #             docling_file = sorted(docling_files)[-1]
        #             try:
        #                 with open(docling_file, 'r', encoding='utf-8') as f:
        #                     docling_text = f.read()
        #                 
        #                 # Versuche Depot-Gebühren aus docling.txt zu extrahieren
        #                 fee_pattern = r'betragen\s+netto\s+([\d,]+)\s+Euro\s+\+\s+(\d+)%\s+USt\s+([\d,]+)\s+Euro\s+=\s+brutto\s+([\d,]+)\s+Euro'
        #                 match = re.search(fee_pattern, docling_text, re.DOTALL | re.IGNORECASE)
        #                 if match:
        #                     result['depot_fee_net'] = float(match.group(1).replace(',', '.'))
        #                     result['vat_rate'] = int(match.group(2))
        #                     result['vat_amount'] = float(match.group(3).replace(',', '.'))
        #                     result['fees'] = float(match.group(4).replace(',', '.'))
        #                     result['is_depot_fee'] = True
        #                     result['is_vat_free'] = False
        #                     print(f"  Info: Depot-Gebühren aus docling.txt extrahiert: {result['fees']} EUR (brutto)")
        #             except Exception as e:
        #                 pass  # Silently ignore errors in fallback
            
        return result
    
    def detect_document_type(self, extracted_text: str) -> str:
        """Erkennt den Dokumenttyp basierend auf Inhalt"""
        text_lower = extracted_text.lower()
        
        # Kosteninformation - diese werden oft fälschlicherweise als Depotabschluss benannt
        if 'information über kosten' in text_lower or 'kosteninformation' in text_lower:
            return 'cost_information'
        
        # Depotabschluss
        if 'depotabschluss' in text_lower or 'ex-post' in text_lower:
            # Prüfe ob es ein echter Depotabschluss mit Bestand ist
            if 'summe kurswerte' in text_lower or 'depotbestand' in text_lower or 'stück' in text_lower:
                return 'depot_statement'
            elif 'kein bestand vorhanden' in text_lower:
                return 'depot_statement_empty'
            else:
                # Zusätzliche Prüfung für versteckte Kosteninformationen
                if 'kosten' in text_lower and 'nebenkosten' in text_lower:
                    return 'cost_information'
                return 'depot_statement'
        
        return 'unknown'
    
    def parse_docling_table(self, docling_path: Path) -> Dict[str, Any]:
        """Extrahiert Tabellendaten aus docling.txt Dateien"""
        result = {
            'isin': None,
            'shares': None,
            'kurswert': None,
            'price_per_share': None,  # Add price per share
            'has_table': False
        }
        
        if not docling_path.exists():
            return result
            
        try:
            with open(docling_path, 'r', encoding='utf-8') as f:
                content = f.read()
                
            # Check for markdown table
            if '|' in content and 'Stück' in content:
                result['has_table'] = True
                lines = content.split('\n')
                
                for i, line in enumerate(lines):
                    # Look for table row with "Stück" and a number
                    if 'Stück' in line and '|' in line:
                        # Try to extract shares from same line
                        shares_match = re.search(r'Stück\s*\|\s*([\d.]+)', line)
                        if shares_match:
                            shares_str = shares_match.group(1).replace('.', '')
                            try:
                                result['shares'] = int(shares_str)
                            except ValueError:
                                pass
                        
                        # Look for ISIN and price info in next line
                        if i + 1 < len(lines):
                            next_line = lines[i + 1]
                            isin_match = re.search(r'([A-Z]{2}[A-Z0-9]{10})', next_line)
                            if isin_match:
                                result['isin'] = isin_match.group(1)
                            
                            # Extract price per share (e.g., "533,70 EUR")
                            # Pattern: number with comma decimal, followed by EUR
                            price_match = re.search(r'([\d]{1,3}(?:\.[\d]{3})*,[\d]{2})\s*EUR', next_line)
                            if price_match:
                                price_str = price_match.group(1).replace('.', '').replace(',', '.')
                                try:
                                    price_val = float(price_str)
                                    # Check if this is likely the price per share (not the total)
                                    # Price per share is typically < 10000
                                    if price_val < 10000:
                                        result['price_per_share'] = price_val
                                except ValueError:
                                    pass
                            
                            # Extract Kurswert (total value - typically larger number)
                            # Look for the larger number that's likely the total
                            all_numbers = re.findall(r'([\d]{1,3}(?:\.[\d]{3})*,[\d]{2})', next_line)
                            if all_numbers:
                                for num_str in all_numbers:
                                    value = float(num_str.replace('.', '').replace(',', '.'))
                                    # Kurswert is typically the larger value
                                    if value > 10000 and (not result['kurswert'] or value > result['kurswert']):
                                        result['kurswert'] = value
                
                # Alternative: Look for ISIN pattern anywhere in content
                if not result['isin']:
                    isin_match = re.search(r'([A-Z]{2}[A-Z0-9]{10})', content)
                    if isin_match:
                        result['isin'] = isin_match.group(1)
                
        except Exception as e:
            print(f"Error parsing docling file {docling_path}: {e}")
            
        return result
    
    def extract_stock_split_data(self, text: str) -> Dict[str, Any]:
        """Extract stock split information from Kapitalmaßnahme documents"""
        split_data = {}
        
        # Parse original shares (Nominale/Stück before split)
        # Pattern: "Nominale ... Stück 200"
        original_match = re.search(r'Nominale.*?Stück\s+(\d+)', text, re.DOTALL | re.IGNORECASE)
        if original_match:
            split_data['original_shares'] = int(original_match.group(1))
        
        # Parse split ratio (Verhältnis: 1 : 2 means 1→3)
        ratio_match = re.search(r'Verhältnis:\s*(\d+)\s*:\s*(\d+)', text, re.IGNORECASE)
        if ratio_match:
            base = int(ratio_match.group(1))
            additional = int(ratio_match.group(2))
            # Calculate multiplier: for 1:2 split, you get 2 additional shares, so total is 3x
            split_data['split_multiplier'] = (base + additional) / base
            split_data['split_ratio_text'] = f"{base}:{additional}"
            
        # Parse new shares added (Einbuchung Stück)
        new_shares_match = re.search(r'Einbuchung\s+Stück\s+(\d+)', text, re.IGNORECASE)
        if new_shares_match:
            split_data['new_shares_added'] = int(new_shares_match.group(1))
            
        # Also check for phrases like "verdreifacht" (tripled)
        if 'dreifache' in text.lower() or 'verdreifacht' in text.lower():
            if 'split_multiplier' not in split_data:
                split_data['split_multiplier'] = 3.0
                
        return split_data
    
    def extract_depot_balance(self, depot_path: Path) -> Dict[str, Any]:
        """Extrahiert Depot-Salden aus Depotabschluss-Dokumenten"""
        depot_statements = []
        cost_information_docs = []  # Separate Liste für Kosteninformationen
        latest_balance = None
        latest_balance_date = None
        
        # Finde alle Depotabschluss-Dateien
        for thea_file in depot_path.glob("*Depotabschluss*.thea_extract"):
            data = self.load_thea_extract(thea_file)
            if data:
                extracted_text = data.get('response', {}).get('json', {}).get('extracted_text', '')
                file_path = data.get('metadata', {}).get('file', {}).get('pdf_path', '')
                file_name = os.path.basename(file_path)
                doc_date = self.extract_date_from_filename(file_name)
                
                # Erkenne Dokumenttyp
                doc_type = self.detect_document_type(extracted_text)
                
                # Überspringe Kosteninformationen für Depot-Tabelle
                if doc_type == 'cost_information':
                    cost_information_docs.append({
                        'doc_date': doc_date,
                        'file': file_name,
                        'pdf_path': file_path,
                        'type': 'cost_information'
                    })
                    continue  # Nicht in Depot-Statements aufnehmen
                
                # Extrahiere das tatsächliche Stichtagsdatum (z.B. "per 31.12.2021")
                balance_date = None
                balance_date_match = re.search(r'per\s+(\d{1,2})\.(\d{1,2})\.(\d{4})', extracted_text)
                if balance_date_match:
                    day, month, year = balance_date_match.groups()
                    balance_date = f"{year}-{month.zfill(2)}-{day.zfill(2)}"
                
                # Bestimme ob Jahres- oder Quartalsabschluss basierend auf Stichtag
                if balance_date:
                    month = int(balance_date[5:7])
                    # Jahresabschlüsse sind immer zum 31.12.
                    statement_type = 'annual' if month == 12 else 'quarterly'
                else:
                    # Fallback auf Dokumentdatum
                    statement_type = 'annual' if doc_date and doc_date[5:7] == '01' else 'quarterly'
                
                # Extrahiere Depotbestand (Closing Balance)
                closing_balance = None
                
                # Extrahiere Stückzahl und ISIN
                shares = None
                isin = None
                security_name = None
                
                # Spezielle Prüfung für "kein Bestand vorhanden"
                if re.search(r'kein\s+Bestand\s+vorhanden', extracted_text, re.IGNORECASE):
                    closing_balance = 0.0
                    shares = None
                    isin = None
                else:
                    # Verbesserte Suche nach Kurswert in Tabellen mit verschiedenen Formatierungen
                    # Spezieller Fall für Tabellenlayout wo "Summe Kurswerte" in einer Spalte und Betrag in anderer ist
                    if 'Summe Kurswerte' in extracted_text or 'Summe' in extracted_text:
                        lines = extracted_text.split('\n')
                        for i, line in enumerate(lines):
                            if 'Summe Kurswerte' in line or ('Summe' in line and '|' in line):
                                # Extrahiere alle Zahlen aus der aktuellen Zeile (deutsche Formatierung mit . als Tausendertrennzeichen)
                                numbers = re.findall(r'[\d]{1,3}(?:\.[\d]{3})*,[\d]{2}', line)
                                # Nimm den größten Wert (der Depotbestand, nicht die Stückzahl)
                                largest_value = 0
                                for num in numbers:
                                    # Konvertiere deutsches in Python-Format
                                    balance_str = num.replace('.', '').replace(',', '.')
                                    try:
                                        test_balance = float(balance_str)
                                        # Nimm den größten Wert, der mindestens 5000 EUR ist
                                        # (um kleine Stückzahlen zu vermeiden, aber erlaubt kleinere Depots)
                                        if test_balance > largest_value and test_balance > 5000:
                                            largest_value = test_balance
                                    except ValueError:
                                        pass
                                
                                # Wenn kein Wert in der gleichen Zeile gefunden, prüfe die nächste Zeile (multi-row table format)
                                if largest_value == 0 and i + 1 < len(lines):
                                    next_line = lines[i + 1]
                                    # Suche nach Werten in der nächsten Zeile (für multi-row tables)
                                    numbers_next = re.findall(r'[\d]{1,3}(?:\.[\d]{3})*,[\d]{2}', next_line)
                                    for num in numbers_next:
                                        balance_str = num.replace('.', '').replace(',', '.')
                                        try:
                                            test_balance = float(balance_str)
                                            # Akzeptiere auch kleinere Werte in multi-row format (min 1000 EUR)
                                            if test_balance > largest_value and test_balance > 1000:
                                                largest_value = test_balance
                                        except ValueError:
                                            pass
                                
                                if largest_value > 0:
                                    closing_balance = largest_value
                                    break
                                if closing_balance and closing_balance > 0:
                                    break
                    
                    # Fallback auf Pattern-Matching
                    if closing_balance is None:
                        kurswert_patterns = [
                            r'Summe\s+Kurswerte\s*\|\s*([+-]?[\d.,]+)',  # Mit Tabellen-Separator
                            r'Summe\s+([+-]?[\d.,]+)',  # Summe gefolgt von Zahl
                            r'([+-]?[\d.,]+)\s*\|\s*$',  # Zahl am Ende einer Tabellenzeile
                            r'Summe\s+Kurswerte\s*([+-]?[\d.,]+)',
                            r'Kurswert\s+in\s+EUR[^\d]*([+-]?[\d.,]+)',  # Kurswert in EUR
                            r'Kurswert.*?([+-]?[\d.,]+)\s*(?:EUR|€)',  # Kurswert mit Betrag
                        ]
                        
                        for pattern in kurswert_patterns:
                            # Suche nach allen Matches, nicht nur dem ersten
                            matches = re.findall(pattern, extracted_text, re.IGNORECASE)
                            if matches:
                                largest_value = 0
                                for match_str in matches:
                                    balance_str = match_str.replace('.', '').replace(',', '.')
                                    try:
                                        test_balance = float(balance_str)
                                        # Nimm den größten Wert über 5000 EUR
                                        if test_balance > largest_value and test_balance > 5000:
                                            largest_value = test_balance
                                    except ValueError:
                                        pass
                                if largest_value > 0:
                                    closing_balance = largest_value
                                    break
                    
                    # Fallback auf andere Balance-Muster
                    if closing_balance is None:
                        closing_patterns = [
                            r'Depotbestand[:\s]+([+-]?[\d.,]+)\s*(?:EUR|€)',
                            r'Durchschnittsdepotbestand[:\s]+([+-]?[\d.,]+)\s*(?:EUR|€)',
                            r'bewerteten?\s+Durchschnittsdepotbestand[:\s]+(?:in Höhe von\s+)?([+-]?[\d.,]+)\s*(?:EUR|Euro|€)',
                            r'Endsaldo[:\s]+([+-]?[\d.,]+)\s*(?:EUR|€)',
                        ]
                        
                        for pattern in closing_patterns:
                            match = re.search(pattern, extracted_text, re.IGNORECASE)
                            if match:
                                balance_str = match.group(1).replace('.', '').replace(',', '.')
                                try:
                                    closing_balance = float(balance_str)
                                    if closing_balance > 0:
                                        break
                                except ValueError:
                                    pass
                
                # Default zu 0 wenn nichts gefunden
                if closing_balance is None:
                    closing_balance = 0.0
                
                # Extrahiere Depot-Gebühren (alle Felder)
                depot_fees = None  # Brutto
                depot_fee_net = None
                vat_rate = None
                vat_amount = None
                
                # Versuche zuerst aus extract_trading_details
                trading_details = self.extract_trading_details(extracted_text)
                if trading_details.get('fees'):
                    depot_fees = trading_details['fees']  # Brutto
                    depot_fee_net = trading_details.get('depot_fee_net')
                    vat_rate = trading_details.get('vat_rate', 19)
                    vat_amount = trading_details.get('vat_amount')
                else:
                    # Fallback auf docling.txt wenn vorhanden
                    import glob
                    if file_path:
                        docling_pattern = f"{file_path}.*docling.txt"
                        docling_files = glob.glob(docling_pattern)
                        if docling_files:
                            try:
                                with open(sorted(docling_files)[-1], 'r', encoding='utf-8') as f:
                                    docling_text = f.read()
                                fee_pattern = r'betragen\s+netto\s+([\d,]+)\s+Euro\s+\+\s+(\d+)%\s+USt\s+([\d,]+)\s+Euro\s+=\s+brutto\s+([\d,]+)\s+Euro'
                                match = re.search(fee_pattern, docling_text, re.DOTALL | re.IGNORECASE)
                                if match:
                                    depot_fee_net = float(match.group(1).replace(',', '.'))
                                    vat_rate = int(match.group(2))
                                    vat_amount = float(match.group(3).replace(',', '.'))
                                    depot_fees = float(match.group(4).replace(',', '.'))
                            except:
                                pass
                
                # Extrahiere Stückzahl und ISIN wenn Bestand vorhanden
                if closing_balance and closing_balance > 0:
                    # Spezialbehandlung für vertikales Layout (Stück auf eigener Zeile)
                    if 'Stück' in extracted_text and shares is None:
                        # Suche nach "Stück" und dann die nächste Zahl
                        lines = extracted_text.split('\n')
                        for i, line in enumerate(lines):
                            if 'Stück' in line:
                                # Schaue in den nächsten 5 Zeilen nach einer Zahl
                                for j in range(i+1, min(i+6, len(lines))):
                                    # Deutsche Zahlenformatierung: 1.300 = 1300
                                    number_match = re.search(r'^\s*([\d.]+)\s*$', lines[j])
                                    if number_match:
                                        number_str = number_match.group(1).replace('.', '')
                                        try:
                                            test_shares = int(number_str)
                                            if 100 <= test_shares <= 10000:  # Plausible Range
                                                shares = test_shares
                                                break
                                        except ValueError:
                                            pass
                                if shares:
                                    break
                    
                    # Fallback auf andere Patterns wenn noch keine Stückzahl gefunden
                    if shares is None:
                        shares_patterns = [
                            # Markdown Tabellen mit | Trennzeichen und deutscher Zahlenformatierung
                            r'\|\s*Stück\s*\|\s*([\d.,]+)',  # | Stück | 1.300 oder | Stück | 1,300
                            # Standard Tabellen mit | Trennzeichen
                            r'\|\s*Stück\s*\|\s*(\d+)',  # | Stück | 950
                            r'Stück\s+\|\s+([\d.,]+)',   # Stück | 1.300
                            r'Stück\s+\|\s+(\d+)',        # Stück | 950
                            
                            # Tabellen ohne Trennzeichen (mit deutscher Zahlenformatierung)
                            r'Stück\s+([\d.]+)(?:\s|$)',     # Stück 950 oder Stück 1.665
                            r'([\d.]+)\s+Stück',              # 950 Stück oder 1.665 Stück
                            
                            # In Zeilen mit TESLA (mit deutscher Zahlenformatierung)
                            r'TESLA[^\n]*?([\d.]{2,6})\s+Stück',
                            r'TESLA[^\n]*?Stück\s+([\d.]{2,6})',
                            
                            # Mit ISIN - removed to avoid capturing dates
                            # r'US88160R1014[^\n]*?(\d{3,4})\s+',
                            r'US88160R1014[^\n]*?Stück\s+([\d.]{2,6})',
                            
                            # Spezielle Formate (mit deutscher Zahlenformatierung)
                            r'Bestand:\s*([\d.]+)\s+Stück',
                            r'Anzahl:\s*(\d+)',
                            r'Position[^\n]*?(\d{3,4})\s+Stück',
                        ]
                        
                        for pattern in shares_patterns:
                            matches = re.findall(pattern, extracted_text)
                            if matches:
                                # Bei mehreren Matches, nimm den größten plausiblen Wert
                                for match in matches:
                                    # Handle German number formatting (1.300 = 1300)
                                    number_str = match.replace('.', '').replace(',', '')
                                    try:
                                        test_shares = int(number_str)
                                        # Exclude year-like numbers (2020-2030) and use plausible share range
                                        if 1 <= test_shares <= 100000 and not (2020 <= test_shares <= 2030):
                                            shares = test_shares
                                            break
                                    except ValueError:
                                        pass
                                if shares:
                                    break
                    
                    # Extrahiere ISIN
                    isin_match = re.search(r'([A-Z]{2}[A-Z0-9]{10})', extracted_text)
                    if isin_match:
                        isin = isin_match.group(1)
                        
                    # Extrahiere Wertpapiername (optional)
                    if isin == 'US88160R1014':
                        security_name = 'TESLA INC'
                
                # Fallback: If no ISIN/shares found in THEA extract, try docling.txt
                # Also try if closing_balance is 0 but it's not explicitly "kein Bestand vorhanden"
                has_no_holdings = re.search(r'kein\s+Bestand\s+vorhanden', extracted_text, re.IGNORECASE)
                if (not isin or not shares) and not has_no_holdings:
                    # Try to find corresponding docling.txt file
                    pdf_base = file_path.replace('.pdf', '')
                    docling_pattern = f"{pdf_base}*.docling.txt"
                    docling_files = list(depot_path.glob(os.path.basename(docling_pattern)))
                    
                    if docling_files:
                        # Use the first matching docling file
                        docling_data = self.parse_docling_table(docling_files[0])
                        
                        if docling_data['has_table']:
                            if not isin and docling_data['isin']:
                                isin = docling_data['isin']
                                print(f"  Found ISIN {isin} from docling table")
                            
                            if not shares and docling_data['shares']:
                                shares = docling_data['shares']
                                print(f"  Found {shares} shares from docling table")
                            
                            # Update closing balance if found in table
                            if docling_data['kurswert'] and docling_data['kurswert'] > 0:
                                closing_balance = docling_data['kurswert']
                                print(f"  Found Kurswert {closing_balance:.2f} EUR from docling table")
                
                # Try to get parsed price per share from docling, calculate as fallback
                price_per_share = None
                price_per_share_parsed = None
                
                # Check if we got a parsed price from docling
                if 'docling_data' in locals() and docling_data.get('price_per_share'):
                    price_per_share_parsed = docling_data['price_per_share']
                    price_per_share = price_per_share_parsed
                    print(f"  Found parsed price per share: {price_per_share:.2f} EUR")
                
                # Calculate price per share as cross-check or fallback
                if shares and shares > 0 and closing_balance and closing_balance > 0:
                    calculated_price = closing_balance / shares
                    
                    # If we have a parsed price, cross-check it
                    if price_per_share_parsed:
                        # Check if they match within 1% tolerance
                        if abs(price_per_share_parsed - calculated_price) / calculated_price > 0.01:
                            print(f"  ⚠️ Price mismatch: Parsed={price_per_share_parsed:.2f}, Calculated={calculated_price:.2f}")
                    else:
                        # Use calculated price if no parsed price available
                        price_per_share = calculated_price
                        print(f"  Calculated price per share: {price_per_share:.2f} EUR")
                
                depot_statements.append({
                    'doc_date': doc_date,
                    'balance_date': balance_date if balance_date else doc_date,
                    'closing_balance': closing_balance,
                    'shares': shares,
                    'isin': isin,
                    'security_name': security_name,
                    'price_per_share': price_per_share,  # Price per share
                    'file': file_name,
                    'type': statement_type,
                    'pdf_path': file_path,
                    'depot_fees': depot_fees,       # Depot-Gebühren (brutto)
                    'depot_fee_net': depot_fee_net, # Depot-Gebühren (netto)
                    'vat_rate': vat_rate,           # USt-Satz
                    'vat_amount': vat_amount        # USt-Betrag
                })
                
                # Aktualisiere letzten Saldo basierend auf Stichtag
                check_date = balance_date if balance_date else doc_date
                if closing_balance > 0:  # Nur positive Salden berücksichtigen
                    if not latest_balance_date or (check_date and check_date > latest_balance_date):
                        latest_balance = closing_balance
                        latest_balance_date = check_date
        
        # Sortiere Statements nach Stichtag
        depot_statements.sort(key=lambda x: x['balance_date'] if x['balance_date'] else '0000-00-00')
        
        # Separiere Jahres- und Quartalsabschlüsse und sortiere nach Stichtag
        annual_statements = [s for s in depot_statements if s['type'] == 'annual']
        annual_statements.sort(key=lambda x: x['balance_date'] if x['balance_date'] else '0000-00-00')
        
        quarterly_statements = [s for s in depot_statements if s['type'] == 'quarterly']
        quarterly_statements.sort(key=lambda x: x['balance_date'] if x['balance_date'] else '0000-00-00')
        
        # Log wenn Kosteninformationen gefunden wurden
        if cost_information_docs:
            print(f"  Info: {len(cost_information_docs)} Kosteninformations-Dokumente gefunden (werden nicht in Depot-Tabelle aufgenommen)")
        
        return {
            'statements': depot_statements,
            'annual_statements': annual_statements,
            'quarterly_statements': quarterly_statements,
            'cost_information_docs': cost_information_docs,  # Füge Kosteninformationen separat hinzu
            'latest_balance': latest_balance,
            'latest_balance_date': latest_balance_date
        }
    
    def extract_cost_information(self, depot_path: Path) -> Dict[str, List[Dict]]:
        """Extrahiert detaillierte Kosteninformationen aus MiFID II Dokumenten"""
        cost_data_by_year = {}
        
        # Finde alle Kosteninformations-Dokumente
        for thea_file in depot_path.glob("*Depotabschluss*.thea_extract"):
            data = self.load_thea_extract(thea_file)
            if data:
                extracted_text = data.get('response', {}).get('json', {}).get('extracted_text', '')
                file_path = data.get('metadata', {}).get('file', {}).get('pdf_path', '')
                file_name = os.path.basename(file_path)
                
                # Prüfe ob es ein Kosteninformations-Dokument ist
                if 'information über kosten' in extracted_text.lower() or 'kosteninformation' in extracted_text.lower():
                    # Extrahiere das Jahr aus dem Text
                    year_match = re.search(r'für das Jahr (\d{4})', extracted_text)
                    if year_match:
                        year = year_match.group(1)
                        
                        cost_info = {
                            'year': year,
                            'doc_date': self.extract_date_from_filename(file_name),
                            'file': file_name,
                            'total_costs': None,
                            'total_costs_net': None,
                            'total_costs_vat': None,
                            'service_costs_once': None,
                            'service_costs_ongoing': None,
                            'depot_fees': None,
                            'depot_fees_net': None,
                            'depot_fees_vat': None,
                            'product_costs': None,
                            'trading_volume': None,
                            'avg_depot_value': None,
                            'cost_percentage_volume': None,
                            'cost_percentage_depot': None,
                            'vat_rate': 19  # Standard USt-Satz in Deutschland
                        }
                        
                        # Extrahiere Dienstleistungskosten (Service/Trading costs)
                        # Handle both formats: with pipe separator and without
                        service_patterns = [
                            r'Dienstleistungskosten\s*\|\s*([\d.,]+)\s*€',  # With pipe separator
                            r'Dienstleistungskosten[^\d]+([\d.,]+)\s*€'     # General pattern
                        ]
                        for pattern in service_patterns:
                            service_match = re.search(pattern, extracted_text)
                            if service_match:
                                cost_info['service_costs'] = float(service_match.group(1).replace('.', '').replace(',', '.'))
                                break
                        else:
                            cost_info['service_costs'] = 0.0
                        
                        # Extrahiere Übergreifende Kosten (Depotentgelte)
                        depot_fees = re.search(r'Übergreifende Kosten[^\d]+([\d.,]+)\s*€', extracted_text)
                        if depot_fees:
                            cost_info['depot_fees'] = float(depot_fees.group(1).replace('.', '').replace(',', '.'))
                        else:
                            cost_info['depot_fees'] = 0.0
                        
                        # Berechne korrekte Gesamtkosten (Service + Depot)
                        cost_info['total_costs'] = cost_info['service_costs'] + cost_info['depot_fees']
                        
                        # Falls keine separaten Kosten gefunden, suche nach Gesamtkosten in der rechten Spalte
                        if cost_info['total_costs'] == 0:
                            # Suche nach dem Pattern "| Gesamtkosten | ... | XXX,XX € |" (rechteste Spalte)
                            total_pattern = r'\|\s*Gesamtkosten\s*\|(?:[^|]*\|){4}\s*([\d.,]+)\s*€'
                            total_match = re.search(total_pattern, extracted_text)
                            if total_match:
                                cost_info['total_costs'] = float(total_match.group(1).replace('.', '').replace(',', '.'))
                        
                        # Extrahiere Umsatzvolumen
                        volume = re.search(r'Umsatzvolumen[^\d]+([\d.,]+)\s*Euro', extracted_text)
                        if volume:
                            cost_info['trading_volume'] = float(volume.group(1).replace('.', '').replace(',', '.'))
                        
                        # Extrahiere durchschnittlichen Depotbestand
                        avg_depot = re.search(r'Durchschnittsdepotbestand[^\d]+([\d.,]+)\s*Euro', extracted_text)
                        if avg_depot:
                            cost_info['avg_depot_value'] = float(avg_depot.group(1).replace('.', '').replace(',', '.'))
                        
                        # Berechne USt-Anteile (die Kosten sind Bruttobeträge inkl. 19% USt)
                        if cost_info['service_costs']:
                            # Nettobetrag = Bruttobetrag / 1.19
                            cost_info['service_costs_net'] = round(cost_info['service_costs'] / 1.19, 2)
                            # USt = Brutto - Netto
                            cost_info['service_costs_vat'] = round(cost_info['service_costs'] - cost_info['service_costs_net'], 2)
                        else:
                            cost_info['service_costs_net'] = 0.0
                            cost_info['service_costs_vat'] = 0.0
                        
                        if cost_info['depot_fees']:
                            # Nettobetrag = Bruttobetrag / 1.19
                            cost_info['depot_fees_net'] = round(cost_info['depot_fees'] / 1.19, 2)
                            # USt = Brutto - Netto  
                            cost_info['depot_fees_vat'] = round(cost_info['depot_fees'] - cost_info['depot_fees_net'], 2)
                        else:
                            cost_info['depot_fees_net'] = 0.0
                            cost_info['depot_fees_vat'] = 0.0
                        
                        # Berechne Gesamt-Netto und Gesamt-USt
                        if cost_info['total_costs']:
                            cost_info['total_costs_net'] = round(cost_info['total_costs'] / 1.19, 2)
                            cost_info['total_costs_vat'] = round(cost_info['total_costs'] - cost_info['total_costs_net'], 2)
                        
                        # Berechne Kostenquoten
                        if cost_info['total_costs'] and cost_info['trading_volume'] and cost_info['trading_volume'] > 0:
                            cost_info['cost_percentage_volume'] = round(cost_info['total_costs'] / cost_info['trading_volume'] * 100, 2)
                        
                        if cost_info['total_costs'] and cost_info['avg_depot_value'] and cost_info['avg_depot_value'] > 0:
                            cost_info['cost_percentage_depot'] = round(cost_info['total_costs'] / cost_info['avg_depot_value'] * 100, 2)
                        
                        cost_data_by_year[year] = cost_info
        
        return cost_data_by_year
    
    def analyze_depot(self, depot_name: str) -> Dict[str, Any]:
        """Analysiert alle Dateien eines Depots mit erweiterter Datenqualitätsprüfung"""
        depot_info = self.depots[depot_name]
        depot_path = self.base_path / depot_info['folder']
        
        if not depot_path.exists():
            print(f"Depot-Ordner nicht gefunden: {depot_path}")
            return None
            
        # Sammle alle .thea_extract Dateien
        thea_files = list(depot_path.glob("*.thea_extract"))
        pdf_files = list(depot_path.glob("*.pdf")) + list(depot_path.glob("*.PDF"))
        
        print(f"Gefunden: {len(pdf_files)} PDF-Dateien und {len(thea_files)} .thea_extract Dateien in {depot_name}")
        
        transactions = []
        statistics = defaultdict(int)
        isin_groups = defaultdict(list)
        
        # Datenqualitäts-Statistiken
        data_quality = {
            'total_transactions': 0,
            'with_fees': 0,
            'without_fees': 0,
            'with_execution_date': 0,
            'with_profit_loss': 0,
            'missing_isin': 0,
            'missing_amounts': 0,
            'suspicious_fees': 0  # Gebühren > 10% des Volumens
        }
        
        for file_path in thea_files:
            data = self.load_thea_extract(file_path)
            if data:
                transaction = self.extract_transaction_data(data)
                if transaction:
                    transactions.append(transaction)
                    statistics[transaction['type']] += 1
                    if transaction['isin']:
                        isin_groups[transaction['isin']].append(transaction)
                    
                    # Sammle Datenqualitäts-Statistiken
                    data_quality['total_transactions'] += 1
                    
                    # Prüfe Gebühren
                    if transaction.get('fees') and transaction['fees'] > 0:
                        data_quality['with_fees'] += 1
                        # Prüfe auf verdächtig hohe Gebühren (> 10% des Volumens)
                        if transaction.get('gross_amount') and transaction['gross_amount'] > 0:
                            fee_percentage = (transaction['fees'] / transaction['gross_amount']) * 100
                            if fee_percentage > 10:
                                data_quality['suspicious_fees'] += 1
                    else:
                        # Prüfe ob legitim keine Gebühren (Kurswert = Ausmachend)
                        if transaction.get('gross_amount') and transaction.get('net_amount'):
                            if abs(transaction['gross_amount'] - transaction['net_amount']) < 0.01:
                                data_quality['without_fees'] += 1  # Legitim keine Gebühren
                    
                    # Prüfe Ausführungsdatum
                    if transaction.get('stichtag') and transaction['stichtag'] != transaction.get('date'):
                        data_quality['with_execution_date'] += 1
                    
                    # Prüfe Gewinn/Verlust bei Verkäufen
                    if 'Verkauf' in transaction['type'] and transaction.get('profit_loss') is not None:
                        data_quality['with_profit_loss'] += 1
                    
                    # Prüfe fehlende ISINs
                    if not transaction.get('isin'):
                        data_quality['missing_isin'] += 1
                    
                    # Prüfe fehlende Beträge
                    if not transaction.get('gross_amount') and not transaction.get('net_amount'):
                        data_quality['missing_amounts'] += 1
        
        # Sortiere Transaktionen mit spezieller Behandlung für Depotabschlüsse
        def get_sort_date(trans):
            # Für Depotabschlüsse: Verwende stichtag (das ist der Bilanzstichtag)
            if 'Depotabschluss' in trans.get('type', ''):
                return trans.get('stichtag') or trans.get('date') or '0000-00-00'
            # Für andere Transaktionen: execution_date mit Fallbacks
            return (trans.get('execution_date') or 
                    trans.get('value_date') or 
                    trans.get('date') or 
                    '0000-00-00')
        
        def get_sort_key(trans):
            # 1. Datum (primäre Sortierung)
            date = get_sort_date(trans)
            
            # 2. Hat Referenznummer? (0 = hat Ref, 1 = keine Ref)
            has_ref = 0 if trans.get('transaction_ref') else 1
            
            # 3. Dokumenttyp-Priorität
            if 'Depotabschluss' in trans.get('type', ''):
                doc_priority = 99  # Immer zuletzt
            elif has_ref == 0:
                doc_priority = 0   # Transaktionen mit Ref zuerst
            else:
                doc_priority = 50  # Andere Dokumente in der Mitte
            
            # 4. Referenznummer für Sortierung (oder hoher Wert wenn keine)
            ref_order = trans.get('transaction_ref', 'ZZZZZZZZ')
            
            return (date, has_ref, doc_priority, ref_order)
        
        transactions.sort(key=get_sort_key)
        
        # Extrahiere Depot-Salden
        balance_info = self.extract_depot_balance(depot_path)
        
        # Erstelle Lookup-Dictionary für Depotabschlüsse nach Dateiname
        depot_statement_lookup = {}
        for stmt in balance_info.get('statements', []):
            file_key = os.path.basename(stmt['file'])
            depot_statement_lookup[file_key] = {
                'isin': stmt.get('isin'),
                'shares': stmt.get('shares'),
                'balance': stmt.get('closing_balance'),
                'price_per_share': stmt.get('price_per_share'),  # Price per share
                'depot_fees': stmt.get('depot_fees'),        # Depot-Gebühren (brutto)
                'depot_fee_net': stmt.get('depot_fee_net'),  # Depot-Gebühren (netto)
                'vat_rate': stmt.get('vat_rate'),           # USt-Satz
                'vat_amount': stmt.get('vat_amount')        # USt-Betrag
            }
        
        # Reichere Transaktionen mit Depotabschluss-Daten an
        for trans in transactions:
            if 'Depotabschluss' in trans.get('type', ''):
                file_key = trans.get('original_file', '')
                if file_key in depot_statement_lookup:
                    depot_data = depot_statement_lookup[file_key]
                    # Übernehme ISIN und Stückzahl aus Depotabschluss
                    if depot_data['isin'] and not trans.get('isin'):
                        trans['isin'] = depot_data['isin']
                    if depot_data['shares'] is not None:
                        trans['shares'] = depot_data['shares']
                    # Übernehme Depot-Gebühren aus Depotabschluss
                    if depot_data.get('depot_fees') is not None and not trans.get('fees'):
                        trans['fees'] = depot_data['depot_fees']  # Brutto
                        trans['depot_fee_net'] = depot_data.get('depot_fee_net')
                        trans['vat_rate'] = depot_data.get('vat_rate')
                        trans['vat_amount'] = depot_data.get('vat_amount')
                        trans['is_depot_fee'] = True
                        trans['is_vat_free'] = False
        
        # Extrahiere Kosteninformationen
        cost_info = self.extract_cost_information(depot_path)
        
        # Ausgabe der Datenqualitäts-Statistiken
        if data_quality['total_transactions'] > 0:
            print(f"\n  Datenqualitäts-Statistiken für {depot_name}:")
            print(f"  - Transaktionen mit Gebühren: {data_quality['with_fees']}/{data_quality['total_transactions']} ({data_quality['with_fees']*100/data_quality['total_transactions']:.1f}%)")
            print(f"  - Transaktionen ohne Gebühren (legitim): {data_quality['without_fees']}/{data_quality['total_transactions']} ({data_quality['without_fees']*100/data_quality['total_transactions']:.1f}%)")
            print(f"  - Mit explizitem Ausführungsdatum: {data_quality['with_execution_date']}/{data_quality['total_transactions']} ({data_quality['with_execution_date']*100/data_quality['total_transactions']:.1f}%)")
            
            # Warnungen für Datenqualitätsprobleme
            if data_quality['suspicious_fees'] > 0:
                print(f"  ⚠️  Verdächtig hohe Gebühren (>10%): {data_quality['suspicious_fees']} Transaktionen")
            if data_quality['missing_isin'] > 0:
                print(f"  ⚠️  Fehlende ISINs: {data_quality['missing_isin']} Transaktionen")
            if data_quality['missing_amounts'] > 0:
                print(f"  ⚠️  Fehlende Beträge: {data_quality['missing_amounts']} Transaktionen")
            
            # Erfolgsrate für Gewinn/Verlust bei Verkäufen (ohne Kapitalmaßnahmen und Ausführungsanzeigen)
            verkauf_count = sum(1 for t in transactions if 'Verkauf' in t['type'] and t['type'] != 'Kapitalmaßnahme' and 'Ausführungsanzeige' not in t['type'])
            if verkauf_count > 0:
                gv_rate = (data_quality['with_profit_loss'] / verkauf_count) * 100
                print(f"  - Verkäufe mit G/V-Daten: {data_quality['with_profit_loss']}/{verkauf_count} ({gv_rate:.1f}%)")
        
        analysis = {
            'depot_name': depot_name,
            'depot_number': depot_info['depot_number'],
            'company_name': depot_info['company_name'],
            'fiscal_year': depot_info.get('fiscal_year', {}),  # Füge Geschäftsjahr-Info hinzu
            'total_pdf_files': len(pdf_files),
            'total_thea_files': len(thea_files),
            'transactions': transactions,
            'statistics': dict(statistics),
            'isin_groups': dict(isin_groups),
            'depot_path': depot_path,
            'depot_info': depot_info,
            'depot_statements': balance_info['statements'],
            'annual_statements': balance_info['annual_statements'],
            'quarterly_statements': balance_info['quarterly_statements'],
            'latest_balance': balance_info['latest_balance'],
            'latest_balance_date': balance_info['latest_balance_date'],
            'cost_information': cost_info,  # Füge Kosteninformationen hinzu
            'data_quality': data_quality  # Füge Datenqualitäts-Statistiken hinzu
        }
        
        # Enrich transactions with cumulative data
        analysis['enriched_transactions'] = self.enrich_transactions_with_cumulative_data(analysis)
        
        return analysis
    
    def enrich_transactions_with_cumulative_data(self, analysis: Dict[str, Any]) -> List[Dict[str, Any]]:
        """
        Reichert Transaktionen mit kumulativen Berechnungen an (G/V, Gebühren, etc.)
        Diese Methode führt die gleichen Berechnungen durch wie in der HTML-Generierung,
        speichert sie aber direkt in den Transaktions-Dictionaries für Excel-Export.
        """
        transactions = analysis.get('transactions', [])
        if not transactions:
            return []
        
        enriched_transactions = []
        fiscal_type = analysis.get('fiscal_year', {}).get('type', 'calendar')
        is_dual_tracking = fiscal_type == 'april_march'
        
        # Kumulative Tracker
        transaction_number = 0
        cumulative_stock_pnl = 0.0
        cumulative_non_stock_pnl = 0.0
        cumulative_fees_net = 0.0
        cumulative_fees_vat = 0.0
        previous_fiscal_year = None
        
        # Calendar year tracking (only when FY != CY)
        if is_dual_tracking:
            cumulative_cy_stock_pnl = 0.0
            cumulative_cy_non_stock_pnl = 0.0
            cumulative_cy_fees_net = 0.0
            cumulative_cy_fees_vat = 0.0
            previous_calendar_year = None
        
        # Depot statement lookup for enrichment
        depot_statement_lookup = {}
        for stmt in analysis.get('depot_statements', []):
            file_key = os.path.basename(stmt['file'])
            depot_statement_lookup[file_key] = {
                'isin': stmt.get('isin'),
                'shares': stmt.get('shares'),
                'balance': stmt.get('closing_balance'),
                'price_per_share': stmt.get('price_per_share'),
                'depot_fees': stmt.get('depot_fees'),
                'depot_fee_net': stmt.get('depot_fee_net'),
                'vat_rate': stmt.get('vat_rate'),
                'vat_amount': stmt.get('vat_amount')
            }
        
        # Track share balances per ISIN
        isin_balances = {}
        
        for idx, trans in enumerate(transactions):
            transaction_number += 1
            enriched = trans.copy()  # Create a copy to avoid modifying original
            
            # Add transaction number
            enriched['nr'] = transaction_number
            
            # Map execution_price to price for Excel compatibility
            if 'execution_price' in trans and trans['execution_price'] is not None:
                enriched['price'] = trans['execution_price']
            elif 'execution_price_avg' in trans and trans['execution_price_avg'] is not None:
                enriched['price'] = trans['execution_price_avg']
            elif not enriched.get('price'):
                enriched['price'] = None
            
            # Date handling
            value_date = trans.get('value_date')
            execution_date = trans.get('execution_date')
            doc_date = trans.get('date')
            
            if not value_date:
                value_date = execution_date or trans.get('stichtag') or doc_date
            if not execution_date:
                execution_date = trans.get('stichtag') or doc_date
            
            enriched['execution_date'] = execution_date
            enriched['value_date'] = value_date
            enriched['doc_date'] = doc_date
            
            # Determine fiscal year
            trans_type = trans.get('type', '')
            if 'Depotabschluss' in trans_type:
                fy_date = trans.get('stichtag') or value_date
            else:
                fy_date = execution_date or value_date
            
            fiscal_year = self.get_fiscal_year(fy_date, fiscal_type)
            enriched['fiscal_year'] = fiscal_year
            
            # Check for fiscal year transition
            if previous_fiscal_year and fiscal_year != previous_fiscal_year:
                cumulative_stock_pnl = 0.0
                cumulative_non_stock_pnl = 0.0
                cumulative_fees_net = 0.0
                cumulative_fees_vat = 0.0
            previous_fiscal_year = fiscal_year
            
            # Calendar year tracking for dual mode
            if is_dual_tracking:
                if 'Depotabschluss' in trans_type:
                    stichtag = trans.get('stichtag')
                    calendar_year = self.get_calendar_year(stichtag) if stichtag else self.get_calendar_year(value_date)
                else:
                    calendar_year = self.get_calendar_year(fy_date)
                enriched['calendar_year'] = calendar_year
                
                # Check for calendar year transition
                if previous_calendar_year and calendar_year != previous_calendar_year:
                    cumulative_cy_stock_pnl = 0.0
                    cumulative_cy_non_stock_pnl = 0.0
                    cumulative_cy_fees_net = 0.0
                    cumulative_cy_fees_vat = 0.0
                previous_calendar_year = calendar_year
            
            # Handle share balance tracking
            trans_isin = trans.get('isin')
            if trans_isin:
                # Update balance tracking
                if 'Depotabschluss' in trans_type:
                    # Depot statement sets the balance
                    if trans.get('shares') is not None:
                        isin_balances[trans_isin] = trans['shares']
                        enriched['balance'] = trans['shares']
                        enriched['shares_change'] = 0
                elif trans.get('shares') is not None:
                    # Regular transaction changes the balance
                    previous_balance = isin_balances.get(trans_isin, 0) or 0
                    shares_change = trans['shares']
                    
                    if 'Kauf' in trans_type:
                        new_balance = previous_balance + shares_change
                        enriched['shares_change'] = shares_change
                    elif 'Verkauf' in trans_type:
                        new_balance = previous_balance - shares_change
                        enriched['shares_change'] = -shares_change
                    else:
                        new_balance = previous_balance
                        enriched['shares_change'] = 0
                    
                    isin_balances[trans_isin] = new_balance
                    enriched['balance'] = new_balance
            
            # Extract and accumulate P&L
            stock_pnl = 0.0
            non_stock_pnl = 0.0
            
            if 'Verkauf' in trans_type and trans.get('profit_loss') is not None:
                # Determine if stock or non-stock based on ISIN
                if trans_isin and self.is_stock_isin(trans_isin):
                    stock_pnl = trans['profit_loss']
                    enriched['stock_pnl'] = stock_pnl
                    enriched['non_stock_pnl'] = 0.0
                else:
                    non_stock_pnl = trans['profit_loss']
                    enriched['non_stock_pnl'] = non_stock_pnl
                    enriched['stock_pnl'] = 0.0
            else:
                enriched['stock_pnl'] = 0.0
                enriched['non_stock_pnl'] = 0.0
            
            # Accumulate P&L
            cumulative_stock_pnl += stock_pnl
            cumulative_non_stock_pnl += non_stock_pnl
            
            # Store FY cumulative values
            enriched['cum_stock_pnl_fy'] = round(cumulative_stock_pnl, 2)
            enriched['cum_non_stock_pnl_fy'] = round(cumulative_non_stock_pnl, 2)
            enriched['cum_total_pnl_fy'] = round(cumulative_stock_pnl + cumulative_non_stock_pnl, 2)
            
            # CY cumulative values (if dual tracking)
            if is_dual_tracking:
                cumulative_cy_stock_pnl += stock_pnl
                cumulative_cy_non_stock_pnl += non_stock_pnl
                enriched['cum_stock_pnl_cy'] = round(cumulative_cy_stock_pnl, 2)
                enriched['cum_non_stock_pnl_cy'] = round(cumulative_cy_non_stock_pnl, 2)
                enriched['cum_total_pnl_cy'] = round(cumulative_cy_stock_pnl + cumulative_cy_non_stock_pnl, 2)
            
            # Extract and accumulate fees
            fees_net = 0.0
            fees_vat = 0.0
            
            if trans.get('fees'):
                fees_gross = trans['fees']
                # Calculate net and VAT
                if trans.get('is_vat_free'):
                    fees_net = fees_gross
                    fees_vat = 0.0
                    enriched['vat_rate'] = 0
                else:
                    # Standard 19% VAT
                    fees_net = round(fees_gross / 1.19, 2)
                    fees_vat = round(fees_gross - fees_net, 2)
                    enriched['vat_rate'] = 19
                
                enriched['fees_net'] = fees_net
                enriched['fees_vat'] = fees_vat
            else:
                enriched['fees_net'] = 0.0
                enriched['fees_vat'] = 0.0
                enriched['vat_rate'] = 0
            
            # Accumulate fees
            cumulative_fees_net += fees_net
            cumulative_fees_vat += fees_vat
            
            # Store FY cumulative fees (use gross for display)
            enriched['cum_fees_fy'] = round(cumulative_fees_net + cumulative_fees_vat, 2)  # Gross cumulative
            enriched['cum_vat_fy'] = round(cumulative_fees_vat, 2)
            
            # CY cumulative fees (if dual tracking)
            if is_dual_tracking:
                cumulative_cy_fees_net += fees_net
                cumulative_cy_fees_vat += fees_vat
                enriched['cum_fees_cy'] = round(cumulative_cy_fees_net + cumulative_cy_fees_vat, 2)  # Gross cumulative
                enriched['cum_vat_cy'] = round(cumulative_cy_fees_vat, 2)
            
            # Calculate derived fields
            enriched['total_pnl'] = stock_pnl + non_stock_pnl
            
            # Add to enriched list
            enriched_transactions.append(enriched)
        
        return enriched_transactions
    
    def _create_error_html(self) -> str:
        """Erstellt eine Fehler-HTML-Seite"""
        return """<!DOCTYPE html>
<html lang="de">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Fehler bei der Analyse</title>
    <style>
        body { font-family: Arial, sans-serif; margin: 40px; background: #f5f5f5; }
        .error { background: #fff; padding: 20px; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }
        h1 { color: #dc3545; }
    </style>
</head>
<body>
    <div class="error">
        <h1>Fehler bei der Analyse</h1>
        <p>Keine Daten verfügbar.</p>
    </div>
</body>
</html>"""
    
    def _create_html_header(self, analysis: Dict[str, Any]) -> str:
        """Erstellt den HTML-Header mit umfassendem CSS"""
        company_name = analysis['company_name']
        depot_number = analysis['depot_number']
        
        return f"""<!DOCTYPE html>
<html lang="de">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{company_name} - Depotkonto {depot_number}</title>
    <style>
        /* Grundlegende Styles */
        body {{
            font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
            margin: 0;
            padding: 0;
            background: white;
            color: #333;
            line-height: 1.6;
            width: 100%;
        }}
        
        /* Überschriften */
        h1 {{
            color: #2c3e50;
            border-bottom: 3px solid #3498db;
            padding: 20px 10px 10px 10px;
            margin: 0 0 30px 0;
        }}
        
        h2 {{
            color: #34495e;
            margin-top: 40px;
            margin-bottom: 20px;
            border-bottom: 2px solid #ecf0f1;
            padding: 0 10px 8px 10px;
        }}
        
        h3 {{
            color: #7f8c8d;
            margin-top: 25px;
            margin-bottom: 15px;
            padding: 0 10px;
        }}
        
        p {{
            padding: 0 10px;
        }}
        
        ul, ol {{
            padding-left: 30px;
            padding-right: 10px;
        }}
        
        /* Info-Box für Kontoübersicht */
        .info-box {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 20px;
            border-radius: 0;
            margin: 0 0 30px 0;
        }}
        
        .info-box .info-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 15px;
            margin-top: 15px;
        }}
        
        .info-box .info-item {{
            background: rgba(255,255,255,0.1);
            padding: 10px;
            border-radius: 4px;
        }}
        
        .info-box .info-label {{
            font-size: 0.85em;
            opacity: 0.9;
            margin-bottom: 5px;
        }}
        
        .info-box .info-value {{
            font-size: 1.2em;
            font-weight: bold;
        }}
        
        /* Tabellen */
        table {{
            width: 100%;
            border-collapse: collapse;
            margin: 20px 0;
            font-size: 14px;
        }}
        
        th {{
            background: #34495e;
            color: white;
            padding: 12px 8px;
            text-align: left;
            font-weight: 500;
            position: sticky;
            top: 0;
            z-index: 10;
        }}
        
        td {{
            padding: 10px 8px;
            border-bottom: 1px solid #ecf0f1;
        }}
        
        /* Fee block visual separation */
        .fee-block-start {{
            border-left: 2px solid #2c3e50 !important;
        }}
        
        .fee-block-end {{
            border-right: 2px solid #2c3e50 !important;
        }}
        
        td.fee-block-start {{
            border-left: 2px solid #dee2e6 !important;
        }}
        
        td.fee-block-end {{
            border-right: 2px solid #dee2e6 !important;
        }}
        
        /* Non-stock block visual separation */
        .non-stock-block-start {{
            border-left: 2px solid #2c3e50 !important;
        }}
        
        .non-stock-block-end {{
            border-right: 2px solid #2c3e50 !important;
        }}
        
        td.non-stock-block-start {{
            border-left: 2px solid #dee2e6 !important;
        }}
        
        td.non-stock-block-end {{
            border-right: 2px solid #dee2e6 !important;
        }}
        
        /* Type column visual separation */
        .type-column {{
            border-left: 2px solid #2c3e50 !important;
            border-right: 2px solid #2c3e50 !important;
        }}
        
        td.type-column {{
            border-left: 1px solid #dee2e6 !important;
            border-right: 1px solid #dee2e6 !important;
        }}
        
        /* Type column borders for colored rows */
        .depot-statement-row td.type-column,
        .profit-row td.type-column,
        .loss-row td.type-column {{
            border-left: 2px solid rgba(255,255,255,0.5) !important;
            border-right: 2px solid rgba(255,255,255,0.5) !important;
        }}
        
        tr:hover {{
            background: #f8f9fa;
        }}
        
        /* Alternating row colors for better readability */
        tbody tr:nth-child(even) {{
            background-color: #f9f9f9;
        }}
        
        /* Prevent line breaks in all table cells */
        td {{
            white-space: nowrap;
        }}
        
        /* Bold styling for cumulative values in year summary rows */
        .bold-value {{
            font-weight: bold;
        }}
        
        /* Spezielle Zeilen-Styles für Transaktionen */
        .profit-row {{
            background-color: #28a745 !important;  /* Grün für Gewinn */
            color: white !important;
        }}
        
        .profit-row td {{
            color: white !important;
        }}
        
        .loss-row {{
            background-color: #dc3545 !important;  /* Rot für Verlust */
            color: white !important;
        }}
        
        .loss-row td {{
            color: white !important;
        }}
        
        .depot-statement-row {{
            background-color: #6c757d !important;  /* Mittleres Grau für normale Depotabschlüsse */
            color: white !important;
        }}
        
        .depot-statement-row td {{
            color: white !important;
        }}
        
        /* Nr. column and Document column should have white background */
        .profit-row td:first-child,
        .loss-row td:first-child,
        .depot-statement-row td:first-child,
        .depot-statement-year-end td:first-child,
        .capital-action-row td:first-child,
        .profit-row td:last-child,
        .loss-row td:last-child,
        .depot-statement-row td:last-child,
        .depot-statement-year-end td:last-child,
        .capital-action-row td:last-child {{
            background-color: white !important;
            color: black !important;
        }}
        
        /* All document links should be black */
        .depot-statement-row a {{
            color: black !important;
            text-decoration: underline !important;
        }}
        
        .depot-statement-year-end {{
            background-color: #000000 !important;  /* Schwarz für Jahresende */
            color: white !important;
        }}
        
        .depot-statement-year-end td {{
            color: white !important;
            border-bottom: none !important;
        }}
        
        .depot-statement-year-end a {{
            color: black !important;
            text-decoration: underline !important;
        }}
        
        .misc-row {{
            color: #6c757d !important;
        }}
        
        .misc-row td {{
            color: #6c757d !important;
        }}
        
        .neutral-row {{
            background-color: #e9ecef !important;
        }}
        
        /* Capital actions (stock splits, etc.) */
        .capital-action-row {{
            background-color: #9c27b0 !important;  /* Purple for stock splits */
            color: white !important;
        }}
        
        .capital-action-row td {{
            color: white !important;
        }}
        
        .capital-action-row a {{
            color: #3498db !important;
            text-decoration: underline !important;
        }}
        
        /* Links */
        a {{
            color: #3498db;
            text-decoration: underline;
        }}
        
        /* Document column links - consistent style */
        td:last-child a {{
            color: black !important;
            text-decoration: underline !important;
        }}
        
        a:hover {{
            text-decoration: underline;
        }}
        
        
        /* Zahlen-Formatierung */
        .number {{
            text-align: right;
            font-family: "Courier New", monospace;
        }}
        
        .positive {{
            color: #27ae60;
            font-weight: bold;
        }}
        
        .negative {{
            color: #e74c3c;
            font-weight: bold;
        }}
        
        /* Statistik-Boxen */
        .stats-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
            gap: 20px;
            margin: 20px 0;
        }}
        
        .stat-card {{
            background: #f8f9fa;
            padding: 15px;
            border-radius: 8px;
            border-left: 4px solid #3498db;
        }}
        
        .stat-card h4 {{
            margin: 0 0 10px 0;
            color: #34495e;
        }}
        
        .stat-card ul {{
            margin: 5px 0;
            padding-left: 20px;
        }}
        
        /* Warnungen */
        .warning {{
            background: #fff3cd;
            border-left: 4px solid #ffc107;
            padding: 15px;
            margin: 20px 0;
            border-radius: 4px;
        }}
        
        .warning strong {{
            color: #856404;
        }}
        
        /* Erläuterungen */
        .explanation-box {{
            background: #e3f2fd;
            border-left: 4px solid #2196f3;
            padding: 15px;
            margin: 20px 0;
            border-radius: 4px;
        }}
        
        /* Jahresübergangs-Saldenzeilen */
        .year-summary-row {{
            background-color: #ffd700 !important;  /* Gold/Gelb */
            color: #000000 !important;
        }}
        
        .year-summary-row td {{
            color: #000000 !important;
        }}
        
        .year-start-row {{
            background-color: #ffeb3b !important;  /* Helleres Gelb */
            font-style: italic;
            color: #000000 !important;
        }}
        
        .year-separator {{
            border-bottom: 3px solid #000000 !important;
        }}
        
        /* Calendar year transitions (blue) - only for FY != CY */
        .calendar-year-summary-row {{
            background-color: #1976d2 !important;  /* Blue */
            color: #ffffff !important;  /* White text */
        }}
        
        .calendar-year-start-row {{
            background-color: #1565c0 !important;  /* Slightly darker blue */
            font-style: italic;
            color: #ffffff !important;  /* White text */
        }}
        
        .calendar-year-separator {{
            border-bottom: 3px solid #000000 !important;  /* Black border like FY separator */
        }}
        
        /* Column coloring for FY and CY columns */
        td.fy-column, .fy-column {{
            background-color: #ffd700 !important;  /* Gold/Gelb - same as year-summary-row */
            color: #000000 !important;  /* Black text */
        }}
        
        td.cy-column, .cy-column {{
            background-color: #1976d2 !important;  /* Blue - same as calendar-year-summary-row */
            color: #ffffff !important;  /* White text */
        }}
        
        /* Row color takes precedence over column color */
        .year-summary-row td.cy-column {{
            background-color: #ffd700 !important;  /* Keep yellow in yellow rows */
            color: #000000 !important;  /* Black text */
        }}
        
        .calendar-year-summary-row td.fy-column {{
            background-color: #1976d2 !important;  /* Keep blue in blue rows */
            color: #ffffff !important;  /* White text */
        }}
        
        .explanation-box h4 {{
            margin-top: 0;
            color: #1565c0;
        }}
        
        /* Footer */
        .footer {{
            margin-top: 50px;
            padding-top: 20px;
            border-top: 2px solid #ecf0f1;
            text-align: center;
            color: #7f8c8d;
            font-size: 0.9em;
        }}
        
        /* Responsive Design */
        @media (max-width: 768px) {{
            table {{
                font-size: 12px;
            }}
            
            th, td {{
                padding: 6px 4px;
            }}
            
            .info-grid {{
                grid-template-columns: 1fr;
            }}
        }}
        
        /* Navigation TOC styles */
        .nav-toc {{
            background: #f8f9fa;
            border: 1px solid #dee2e6;
            border-radius: 8px;
            padding: 20px;
            margin: 20px 0;
            position: sticky;
            top: 10px;
            z-index: 5;
        }}
        
        .nav-toc h2 {{
            margin-top: 0;
            color: #2c3e50;
            font-size: 1.2em;
        }}
        
        .nav-toc ul {{
            list-style: none;
            padding-left: 0;
        }}
        
        .nav-toc li {{
            margin: 8px 0;
        }}
        
        .nav-toc a {{
            color: #3498db;
            text-decoration: none;
            padding: 5px 10px;
            display: block;
            border-radius: 4px;
            transition: background 0.3s;
        }}
        
        .nav-toc a:hover {{
            background: #e9ecef;
        }}
        
        /* Toggle button for MiFID section */
        .toggle-btn {{
            background: #6c757d;
            color: white;
            border: none;
            padding: 10px 20px;
            border-radius: 4px;
            cursor: pointer;
            margin: 10px 0;
            font-size: 14px;
        }}
        
        .toggle-btn:hover {{
            background: #5a6268;
        }}
        
        /* Hidden section */
        .hidden-section {{
            display: none;
        }}
        
        /* Print styles */
        @media print {{
            body {{
                background: white;
            }}
            
            .page-break {{
                page-break-after: always;
            }}
            
            .nav-toc {{
                display: none;
            }}
        }}
    </style>
    <script>
        function toggleMiFID() {{
            const section = document.getElementById('mifid-section');
            const btn = document.getElementById('mifid-toggle');
            if (section.style.display === 'none' || section.style.display === '') {{
                section.style.display = 'block';
                btn.textContent = 'MiFID-Kosten ausblenden';
            }} else {{
                section.style.display = 'none';
                btn.textContent = 'MiFID-Kosten anzeigen';
            }}
        }}
        
        // Smooth scroll for navigation
        document.addEventListener('DOMContentLoaded', function() {{
            document.querySelectorAll('a[href^="#"]').forEach(anchor => {{
                anchor.addEventListener('click', function(e) {{
                    e.preventDefault();
                    const target = document.querySelector(this.getAttribute('href'));
                    if (target) {{
                        target.scrollIntoView({{
                            behavior: 'smooth',
                            block: 'start'
                        }});
                    }}
                }});
            }});
        }});
    </script>
</head>
<body>
"""
    
    def _create_html_footer(self) -> str:
        """Erstellt den HTML-Footer"""
        from datetime import datetime
        current_date = datetime.now().strftime("%d.%m.%Y %H:%M")
        
        return f"""
    <div class="footer">
        <p>Erstellt am {current_date} mit THEA Document Analysis System</p>
    </div>
</body>
</html>"""

    def generate_html(self, analysis: Dict[str, Any]) -> str:
        """Generiert vollständigen HTML-Bericht aus der Analyse"""
        if not analysis:
            return self._create_error_html()
        
        # Symbol for empty/missing values (middle dot instead of dash to avoid confusion with negative values)
        EMPTY_VALUE = '•'  # Unicode U+2022 (bullet)
        
        # Start HTML document
        html = self._create_html_header(analysis)
        
        # Main heading and info box
        html += f"""
    <h1>{analysis['company_name']} - Depotkonto {analysis['depot_number']}</h1>
    
    <div class="info-box">
        <h2 style="color: white; border: none; margin: 0;">Kontoübersicht</h2>
        <div class="info-grid">
            <div class="info-item">
                <div class="info-label">Kontoinhaber</div>
                <div class="info-value">{analysis['company_name']}</div>
            </div>
            <div class="info-item">
                <div class="info-label">Depotnummer</div>
                <div class="info-value">{analysis['depot_number']}</div>
            </div>
            <div class="info-item">
                <div class="info-label">PDF-Dokumente</div>
                <div class="info-value">{analysis['total_pdf_files']}</div>
            </div>
            <div class="info-item">
                <div class="info-label">THEA-Analysen</div>
                <div class="info-value">{analysis['total_thea_files']}</div>
            </div>"""
        
        # Füge Geschäftsjahr-Information hinzu
        if analysis.get('fiscal_year'):
            fiscal_year = analysis['fiscal_year']
            html += f"""
            <div class="info-item">
                <div class="info-label">Geschäftsjahr</div>
                <div class="info-value">{fiscal_year.get('description', 'Nicht definiert')}</div>
            </div>"""
        
        if analysis.get('latest_balance'):
            latest_balance_str = self.format_number_german(analysis['latest_balance'], 2)
            latest_date = self.format_date_german(analysis.get('latest_balance_date'))
            html += f"""
            <div class="info-item">
                <div class="info-label">Letzter Depotbestand</div>
                <div class="info-value">{latest_balance_str} EUR</div>
            </div>
            <div class="info-item">
                <div class="info-label">Stand</div>
                <div class="info-value">{latest_date}</div>
            </div>"""
        
        html += """
        </div>
    </div>
"""
        
        # SECTION 1: Depotabschluss-Übersicht (Depot Statement Overview)
        if analysis.get('depot_statements'):
            html += "<h2>Depotabschluss-Übersicht</h2>"
            
            # Combine annual and quarterly statements
            all_statements = []
            
            for statement in analysis.get('annual_statements', []):
                statement_copy = statement.copy()
                statement_copy['type_display'] = 'Jahresabschluss'
                all_statements.append(statement_copy)
            
            for statement in analysis.get('quarterly_statements', []):
                statement_copy = statement.copy()
                statement_copy['type_display'] = 'Quartalsabschluss'
                all_statements.append(statement_copy)
            
            # Sort by balance date
            all_statements.sort(key=lambda x: x['balance_date'] if x['balance_date'] else '0000-00-00')
            
            if all_statements:
                html += """
    <h3>Alle Depotabschlüsse (chronologisch)</h3>
    <table>
        <thead>
            <tr>
                <th>Typ</th>
                <th>Stichtag</th>
                <th>Dokumentdatum</th>
                <th>Stück</th>
                <th>ISIN</th>
                <th>Kurs (EUR)</th>
                <th>Depotbestand (EUR)</th>
                <th>Gebühren netto</th>
                <th>USt %</th>
                <th>USt (EUR)</th>
                <th>Gebühren brutto</th>
                <th>Dokument</th>
            </tr>
        </thead>
        <tbody>
"""
                for statement in all_statements:
                    type_display = statement['type_display']
                    doc_date = self.format_date_german(statement.get('doc_date', 'N/A'))
                    balance_date = self.format_date_german(statement.get('balance_date', 'N/A'))
                    shares_str = str(statement['shares']) if statement.get('shares') else '-'
                    isin_str = statement.get('isin', '-')
                    
                    if statement.get('closing_balance') is not None:
                        closing = self.format_number_german(statement['closing_balance'], 2)
                    else:
                        closing = 'N/A'
                    
                    file_name = statement['file']
                    display_name = file_name[:37] + '...' if len(file_name) > 40 else file_name
                    doc_link = f'<a href="docs/{analysis["depot_info"]["folder"]}/{file_name}">{display_name}</a>'
                    
                    # Format depot fees
                    depot_fee_net_str = EMPTY_VALUE
                    vat_rate_str = EMPTY_VALUE
                    vat_amount_str = EMPTY_VALUE
                    depot_fees_str = EMPTY_VALUE
                    
                    if statement.get('depot_fee_net') is not None:
                        depot_fee_net_str = self.format_number_german(statement['depot_fee_net'], 2)
                    if statement.get('vat_rate') is not None:
                        vat_rate_str = f"{statement['vat_rate']}%"
                    if statement.get('vat_amount') is not None:
                        vat_amount_str = self.format_number_german(statement['vat_amount'], 2)
                    if statement.get('depot_fees') is not None:
                        depot_fees_str = self.format_number_german(statement['depot_fees'], 2)
                    
                    # Format price per share
                    price_per_share_str = EMPTY_VALUE
                    if statement.get('price_per_share') is not None:
                        price_per_share_str = self.format_number_german(statement['price_per_share'], 2)
                    
                    html += f"""
            <tr>
                <td>{type_display}</td>
                <td>{balance_date}</td>
                <td>{doc_date}</td>
                <td class="number">{shares_str}</td>
                <td>{isin_str}</td>
                <td class="number">{price_per_share_str}</td>
                <td class="number">{closing}</td>
                <td class="number">{depot_fee_net_str}</td>
                <td class="number">{vat_rate_str}</td>
                <td class="number">{vat_amount_str}</td>
                <td class="number">{depot_fees_str}</td>
                <td>{doc_link}</td>
            </tr>"""
                
                html += """
        </tbody>
    </table>
"""
                # Summary
                annual_count = len(analysis.get('annual_statements', []))
                quarterly_count = len(analysis.get('quarterly_statements', []))
                html += f'<p><em>Gesamt: {annual_count} Jahresabschlüsse und {quarterly_count} Quartalsabschlüsse</em></p>'
            
            # Overall summary
            if analysis.get('latest_balance'):
                latest_date = self.format_date_german(analysis.get('latest_balance_date', 'N/A'))
                latest_balance_str = self.format_number_german(analysis['latest_balance'], 2)
                html += f'<p><strong>Letzter bekannter Depotbestand: {latest_balance_str} EUR</strong> (Stand: {latest_date})</p>'
            
            total_statements = len(analysis.get('depot_statements', []))
            if total_statements > 0:
                first_date = min(s['balance_date'] for s in analysis['depot_statements'] if s.get('balance_date'))
                last_date = max(s['balance_date'] for s in analysis['depot_statements'] if s.get('balance_date'))
                first_date_german = self.format_date_german(first_date)
                last_date_german = self.format_date_german(last_date)
                html += f'<p><em>Gesamtzeitraum: {first_date_german} bis {last_date_german}</em></p>'
        
        # SECTION 2: Jährliche Kostenanalyse (MiFID II)
        if analysis.get('cost_information'):
            html += """
    <button id="mifid-toggle" class="toggle-btn" style="display: none;" onclick="toggleMiFID()">MiFID-Kosten anzeigen</button>
    <div id="mifid-section" class="hidden-section" style="display: none;">
    <h2 id="mifid-costs">Jährliche Kostenanalyse (MiFID II)</h2>
    <p>Gesetzlich vorgeschriebene Kostenaufstellung gemäß Art. 50 der Verordnung (EU) 2017/565:</p>
    <table>
        <thead>
            <tr>
                <th>Jahr</th>
                <th>Dokumentdatum</th>
                <th>Dienstleistungskosten (netto)</th>
                <th>Depotentgelte (netto)</th>
                <th>Zwischensumme (netto)</th>
                <th>USt (19%)</th>
                <th>Gesamtkosten (brutto)</th>
                <th>Umsatzvolumen</th>
                <th>Ø Depotbestand</th>
                <th>Kostenquote</th>
                <th>Dokument</th>
            </tr>
        </thead>
        <tbody>
"""
            # Sort by year
            years = sorted(analysis['cost_information'].keys())
            
            for year in years:
                cost_data = analysis['cost_information'][year]
                
                # Format amounts in German style
                service_costs_net = self.format_number_german(cost_data.get('service_costs_net', 0), 2)
                depot_fees_net = self.format_number_german(cost_data.get('depot_fees_net', 0), 2)
                subtotal_net = self.format_number_german(cost_data.get('total_costs_net', 0), 2)
                total_costs = self.format_number_german(cost_data.get('total_costs', 0), 2)
                total_vat = self.format_number_german(cost_data.get('total_costs_vat', 0), 2)
                volume = self.format_number_german(cost_data.get('trading_volume', 0), 2) if cost_data.get('trading_volume') else '-'
                avg_depot = self.format_number_german(cost_data.get('avg_depot_value', 0), 2) if cost_data.get('avg_depot_value') else '-'
                
                # Cost percentage
                if cost_data.get('cost_percentage_depot'):
                    cost_quote = self.format_number_german(cost_data['cost_percentage_depot'], 2) + '%'
                else:
                    cost_quote = '-'
                
                # Format document date
                doc_date = self.format_date_german(cost_data.get('doc_date', ''))
                
                # Create document link
                file_name = cost_data.get('file', '')
                display_name = file_name[:37] + "..." if len(file_name) > 40 else file_name
                doc_link = f'<a href="docs/{analysis["depot_info"]["folder"]}/{file_name}">{display_name}</a>'
                
                html += f"""
            <tr>
                <td>{year}</td>
                <td>{doc_date}</td>
                <td class="number">{service_costs_net}</td>
                <td class="number">{depot_fees_net}</td>
                <td class="number">{subtotal_net}</td>
                <td class="number">{total_vat}</td>
                <td class="number"><strong>{total_costs}</strong></td>
                <td class="number">{volume}</td>
                <td class="number">{avg_depot}</td>
                <td class="number">{cost_quote}</td>
                <td>{doc_link}</td>
            </tr>"""
            
            html += """
        </tbody>
    </table>
"""
            # Cost summary
            if len(years) > 0:
                total_all_costs = sum(c.get('total_costs', 0) for c in analysis['cost_information'].values())
                total_all_volume = sum(c.get('trading_volume', 0) for c in analysis['cost_information'].values() if c.get('trading_volume'))
                
                if total_all_costs > 0:
                    total_str = self.format_number_german(total_all_costs, 2)
                    html += f'<p><strong>Gesamtkosten {years[0]}-{years[-1]}: {total_str} €</strong></p>'
                
                if total_all_volume > 0:
                    volume_str = self.format_number_german(total_all_volume, 2)
                    html += f'<p><em>Gesamtes Handelsvolumen: {volume_str} €</em></p>'
                
                # Explanations
                html += """
    <div class="explanation-box">
        <h3>Erläuterungen zur Kostenberechnung:</h3>
        <h4>Kostenzusammensetzung:</h4>
        <ul>
            <li><strong>Dienstleistungskosten:</strong> Beinhalten alle Kosten für Wertpapiergeschäfte (Kauf/Verkauf), Ordergebühren und sonstige Transaktionskosten</li>
            <li><strong>Depotentgelte:</strong> Jährliche Verwahrungsgebühren für die Führung des Depotkontos (übergreifende Kosten)</li>
            <li><strong>Zwischensumme (netto):</strong> Summe aus Dienstleistungskosten und Depotentgelten ohne Umsatzsteuer</li>
            <li><strong>USt (19%):</strong> Gesetzliche Umsatzsteuer auf alle Kosten</li>
            <li><strong>Gesamtkosten (brutto):</strong> Vollständige Kosten inklusive 19% Umsatzsteuer</li>
        </ul>
        <h4>Wichtige Hinweise:</h4>
        <ul>
            <li>✓ Die Depotentgelte sind <strong>in den Gesamtkosten enthalten</strong></li>
            <li>✓ Alle Beträge in den verlinkten Dokumenten sind <strong>Bruttobeträge</strong> (bereits inkl. 19% USt)</li>
            <li>✓ Die Nettobeträge werden berechnet: Netto = Brutto ÷ 1,19</li>
        </ul>
        <p><em>Diese Aufstellung entspricht den gesetzlichen Anforderungen gemäß Art. 50 der Verordnung (EU) 2017/565 (MiFID II).</em></p>
    </div>
    </div>
"""
        
        # SECTION 3: Transaction statistics
        if analysis.get('statistics'):
            html += "<h2>Transaktionsstatistik</h2>"
            html += '<div class="stats-grid">'
            
            # Group statistics
            categories = {
                'Depotabschluss': [],
                'Orderabrechnung': [],
                'Kostenaufstellung': [],
                'Sonstige': []
            }
            
            for trans_type, count in sorted(analysis['statistics'].items()):
                if 'Depotabschluss' in trans_type:
                    categories['Depotabschluss'].append((trans_type, count))
                elif 'Orderabrechnung' in trans_type:
                    categories['Orderabrechnung'].append((trans_type, count))
                elif 'Kostenaufstellung' in trans_type:
                    categories['Kostenaufstellung'].append((trans_type, count))
                else:
                    categories['Sonstige'].append((trans_type, count))
            
            for category, items in categories.items():
                if items:
                    html += f'<div class="stat-card"><h4>{category}</h4><ul>'
                    for trans_type, count in items:
                        display_type = trans_type.replace(f'{category}-', '')
                        html += f'<li>{display_type}: {count}</li>'
                    html += '</ul></div>'
            
            html += '</div>'
        
        # SECTION 4: Wertpapiere nach ISIN
        if analysis.get('isin_groups'):
            html += "<h2>Wertpapiere (nach ISIN)</h2>"
            for isin, trans_list in analysis['isin_groups'].items():
                if isin == "DE0001234567":
                    html += f'<h3>⚠️ {isin} (UNGÜLTIGE TEST-ISIN)</h3>'
                    html += '<p class="warning">⚠️ <strong>WARNUNG: Dies ist eine Test-ISIN und sollte nicht in Produktivdaten vorkommen!</strong></p>'
                else:
                    html += f'<h3>{isin}</h3>'
                
                html += f'<p><strong>Anzahl Transaktionen:</strong> {len(trans_list)}</p>'
                
                # Calculate sums
                verkauf_summe = sum(t['max_amount'] for t in trans_list if t['type'] == 'Verkauf' or 'Verkauf' in t['type'])
                kauf_summe = sum(t['max_amount'] for t in trans_list if t['type'] == 'Kauf' or 'Kauf' in t['type'])
                
                if verkauf_summe > 0:
                    html += f'<p><strong>Verkaufssumme:</strong> {self.format_number_german(verkauf_summe, 2)} EUR</p>'
                if kauf_summe > 0:
                    html += f'<p><strong>Kaufsumme:</strong> {self.format_number_german(kauf_summe, 2)} EUR</p>'
        
        # SECTION 5: Enhanced Transaction table with cumulative P&L
        if analysis.get('transactions'):
            # Get fiscal type early for table structure decisions
            fiscal_type = analysis.get('fiscal_year', {}).get('type', 'calendar')
            
            html += """
    <h2>Transaktionsverlauf (Detailliert)</h2>
    
    <h3>Legende der Spalten:</h3>
    <ul>
        <li><strong>Nr.:</strong> Laufende Nummer der Transaktion</li>
        <li><strong>Transaktion:</strong> Transaktionsdatum/Handelstag (Ausführung)</li>
        <li><strong>Referenz:</strong> Auftragsnummer/Referenznummer</li>
        <li><strong>Wertst.:</strong> Wertstellungsdatum/Valuta (Settlement)</li>
        <li><strong>Dokument:</strong> Dokumenterstellungsdatum</li>
        <li><strong>Gebühren:</strong> Nettogebühren ohne Umsatzsteuer (EUR)</li>
        <li><strong>USt %:</strong> Umsatzsteuersatz (0% für umsatzsteuerfreie Gebühren, 19% für normale Gebühren)</li>
        <li><strong>USt:</strong> Umsatzsteuerbetrag (EUR)</li>
        <li><strong>Geb. Brutto:</strong> Bruttogebühren inklusive Umsatzsteuer (EUR)</li>
        <li><strong>Ausmach.:</strong> Ausmachender Betrag = Kurswert + Gebühren Brutto (EUR)</li>"""
            
            # Add FY/CY specific legends based on fiscal type
            if fiscal_type == 'april_march':
                html += """
        <li><strong>Kum. Geb. (FY):</strong> Kumulierte Nettogebühren im Geschäftsjahr (EUR)</li>
        <li><strong>Kum. Geb. (CY):</strong> Kumulierte Nettogebühren im Kalenderjahr (EUR)</li>
        <li><strong>Kum. USt (FY):</strong> Kumulierte Umsatzsteuer im Geschäftsjahr (EUR)</li>
        <li><strong>Kum. USt (CY):</strong> Kumulierte Umsatzsteuer im Kalenderjahr (EUR)</li>
        <li><strong>G/V Akt.:</strong> Gewinn/Verlust aus Aktienverkäufen (EUR)</li>
        <li><strong>Kum. Akt. (FY):</strong> Kumulierter Gewinn/Verlust Aktien im Geschäftsjahr (EUR)</li>
        <li><strong>Kum. Akt. (CY):</strong> Kumulierter Gewinn/Verlust Aktien im Kalenderjahr (EUR)</li>
        <li><strong>G/V N-Akt.:</strong> Gewinn/Verlust aus Nicht-Aktien (Derivate, Zertifikate, strukturierte Produkte) (EUR)</li>
        <li><strong>Kum. N-Akt. (FY):</strong> Kumulierter Gewinn/Verlust Nicht-Aktien im Geschäftsjahr (EUR)</li>
        <li><strong>Kum. N-Akt. (CY):</strong> Kumulierter Gewinn/Verlust Nicht-Aktien im Kalenderjahr (EUR)</li>
"""
            else:
                html += """
        <li><strong>Kum. Geb.:</strong> Kumulierte Nettogebühren (EUR)</li>
        <li><strong>Kum. USt:</strong> Kumulierte Umsatzsteuer (EUR)</li>
        <li><strong>G/V Akt.:</strong> Gewinn/Verlust aus Aktienverkäufen (EUR)</li>
        <li><strong>Kum. Akt.:</strong> Kumulierter Gewinn/Verlust Aktien (EUR)</li>
        <li><strong>G/V N-Akt.:</strong> Gewinn/Verlust aus Nicht-Aktien (Derivate, Zertifikate, strukturierte Produkte) (EUR)</li>
        <li><strong>Kum. N-Akt.:</strong> Kumulierter Gewinn/Verlust Nicht-Aktien (EUR)</li>
"""
            
            html += """
    </ul>
    
    <table>
        <thead>
            <tr>
                <th>Nr.</th>
                <th>Transaktion</th>
                <th>Referenz</th>
                <th>Wertst.</th>
                <th>Dokument</th>
                <th class="type-column">Typ</th>
                <th>ISIN</th>
                <th>Bestand</th>
                <th>Veränderung</th>
                <th>Kurs</th>
                <th>Kurswert</th>
                <th class="fee-block-start">Gebühren</th>
                <th>USt %</th>
                <th>USt</th>
                <th>Geb. Brutto</th>
                <th>Ausmach.</th>"""
            
            # Add cumulative column headers based on fiscal type
            if fiscal_type == 'april_march':
                html += """
                <th>Kum. Geb. (FY)</th>
                <th>Kum. Geb. (CY)</th>
                <th>Kum. USt (FY)</th>
                <th class="fee-block-end">Kum. USt (CY)</th>
                <th>G/V Akt.</th>
                <th>Kum. Akt. (FY)</th>
                <th>Kum. Akt. (CY)</th>
                <th class="non-stock-block-start">G/V N-Akt.</th>
                <th>Kum. N-Akt. (FY)</th>
                <th class="non-stock-block-end">Kum. N-Akt. (CY)</th>"""
            else:
                html += """
                <th>Kum. Geb.</th>
                <th class="fee-block-end">Kum. USt</th>
                <th>G/V Akt.</th>
                <th>Kum. Akt.</th>
                <th class="non-stock-block-start">G/V N-Akt.</th>
                <th class="non-stock-block-end">Kum. N-Akt.</th>"""
                
            html += """
                <th>Dokument</th>
            </tr>
        </thead>
        <tbody>
"""
            
            # Create depot statement lookup for enrichment
            depot_statement_lookup = {}
            for stmt in analysis.get('depot_statements', []):
                file_key = os.path.basename(stmt['file'])
                depot_statement_lookup[file_key] = {
                    'isin': stmt.get('isin'),
                    'shares': stmt.get('shares'),
                    'balance': stmt.get('closing_balance'),
                    'price_per_share': stmt.get('price_per_share'),  # Price per share
                    'depot_fees': stmt.get('depot_fees'),        # Depot-Gebühren (brutto)
                    'depot_fee_net': stmt.get('depot_fee_net'),  # Depot-Gebühren (netto)
                    'vat_rate': stmt.get('vat_rate'),           # USt-Satz
                    'vat_amount': stmt.get('vat_amount')        # USt-Betrag
                }
            
            # Symbol for empty/missing values (middle dot instead of dash to avoid confusion with negative values)
            EMPTY_VALUE = '•'  # Unicode U+2022 (bullet)
            
            transaction_number = 0
            cumulative_stock_pnl = 0.0
            cumulative_non_stock_pnl = 0.0
            cumulative_total_pnl = 0.0
            cumulative_fees_net = 0.0
            cumulative_fees_vat = 0.0
            
            # Tracking für alle Geschäftsjahre (wird nie zurückgesetzt)
            all_years_stock_pnl = 0.0
            all_years_non_stock_pnl = 0.0
            all_years_total_pnl = 0.0
            all_years_fees_net = 0.0
            all_years_fees_vat = 0.0
            
            # Track transaction numbers per fiscal year
            fiscal_year_counters = {}
            # Track summaries per fiscal year for detailed overview
            fiscal_year_summaries = {}
            
            # Calendar year tracking (only used when FY != CY)
            is_dual_tracking = fiscal_type == 'april_march'  # FY != CY
            if is_dual_tracking:
                cumulative_cy_stock_pnl = 0.0
                cumulative_cy_non_stock_pnl = 0.0
                cumulative_cy_total_pnl = 0.0
                cumulative_cy_fees_net = 0.0
                cumulative_cy_fees_vat = 0.0
                calendar_year_counters = {}
                calendar_year_summaries = {}
                previous_calendar_year = None  # Track previous CY to detect transitions
            
            # Track current share balances per ISIN
            # None = unknown (before first depot statement), number = known balance
            isin_balances = {}
            depot_statement_seen = False  # Track if we've seen any depot statement
            
            for idx, trans in enumerate(analysis['transactions']):
                transaction_number += 1
                
                # Date formatting
                value_date = trans.get('value_date')
                execution_date = trans.get('execution_date')
                doc_date = trans.get('date')
                
                if not value_date:
                    value_date = execution_date or trans.get('stichtag') or doc_date
                if not execution_date:
                    execution_date = trans.get('stichtag') or doc_date
                
                trans_type = trans['type']
                
                # Determine fiscal year - use execution date for regular transactions, stichtag for Depotabschluss
                # This ensures transactions are assigned to the FY when they were executed, not when they settle
                if 'Depotabschluss' in trans_type:
                    fy_date = trans.get('stichtag') or value_date
                else:
                    # For regular transactions, use execution_date (when trade happened)
                    # not value_date (when money settles, which can be in next FY)
                    fy_date = execution_date or value_date
                fiscal_year = self.get_fiscal_year(fy_date, fiscal_type)
                trans['fiscal_year'] = fiscal_year  # Store fiscal_year in transaction for later filtering
                
                # Track calendar year for dual tracking
                if is_dual_tracking:
                    if 'Depotabschluss' in trans_type:
                        # For Depotabschluss, use stichtag for calendar year (the actual reporting date)
                        # This ensures Dec 31 Depotabschluss is assigned to the correct calendar year
                        # Use stichtag directly, not the fallback value_date which might be a later date
                        stichtag = trans.get('stichtag')
                        calendar_year = self.get_calendar_year(stichtag) if stichtag else self.get_calendar_year(value_date)
                    else:
                        calendar_year = self.get_calendar_year(fy_date)
                    trans['calendar_year'] = calendar_year
                    
                    # Check if we've entered a new calendar year (reset CY cumulative values)
                    if previous_calendar_year and calendar_year != previous_calendar_year:
                        # We've crossed into a new calendar year - reset CY cumulative values
                        cumulative_cy_stock_pnl = 0.0
                        cumulative_cy_non_stock_pnl = 0.0
                        cumulative_cy_total_pnl = 0.0
                        cumulative_cy_fees_net = 0.0
                        cumulative_cy_fees_vat = 0.0
                    
                    # Get or initialize counter for this calendar year
                    if calendar_year not in calendar_year_counters:
                        calendar_year_counters[calendar_year] = 0
                    calendar_year_counters[calendar_year] += 1
                
                # Check if this depot statement marks the end of a fiscal year
                is_year_end_depot = False
                if 'Depotabschluss' in trans_type:
                    # Check if the next transaction has a different fiscal year
                    if idx + 1 < len(analysis['transactions']):
                        next_trans = analysis['transactions'][idx + 1]
                        # Determine FY of next transaction
                        next_date = next_trans.get('value_date') or next_trans.get('execution_date') or next_trans.get('stichtag') or next_trans.get('date')
                        if 'Depotabschluss' in next_trans.get('type', ''):
                            next_fy_date = next_trans.get('stichtag') or next_date
                        else:
                            next_fy_date = next_date
                        next_fiscal_year = self.get_fiscal_year(next_fy_date, fiscal_type) if next_fy_date else fiscal_year
                        
                        if next_fiscal_year != fiscal_year:
                            is_year_end_depot = True
                    else:
                        # Last transaction is always year-end if it's a Depotabschluss
                        is_year_end_depot = True
                
                # Check for calendar year transition (only when dual tracking)
                # Only insert separator if this is truly the LAST transaction of the calendar year
                is_calendar_year_end = False
                next_calendar_year = None
                if is_dual_tracking and idx + 1 < len(analysis['transactions']):
                    next_trans = analysis['transactions'][idx + 1]
                    next_date = next_trans.get('value_date') or next_trans.get('execution_date') or next_trans.get('stichtag') or next_trans.get('date')
                    if next_date:
                        current_cy = trans.get('calendar_year', '')
                        next_cy = self.get_calendar_year(next_date)
                        if current_cy and next_cy and current_cy != next_cy:
                            # Check if we're at a significant year boundary
                            # Only insert separator if this is a Depotabschluss OR if it's actually Dec 31
                            if 'Depotabschluss' in trans_type:
                                # For Depotabschluss at calendar year boundary
                                # The stichtag (reporting date) determines if it's a year-end report
                                stichtag = trans.get('stichtag')
                                check_date = stichtag or value_date
                                
                                if check_date:
                                    import datetime
                                    if isinstance(check_date, str):
                                        try:
                                            if '-' in check_date:
                                                date_obj = datetime.datetime.strptime(check_date, '%Y-%m-%d')
                                            elif '.' in check_date:
                                                date_obj = datetime.datetime.strptime(check_date, '%d.%m.%Y')
                                            else:
                                                date_obj = None
                                        except:
                                            date_obj = None
                                    else:
                                        date_obj = check_date
                                    
                                    if date_obj:
                                        # Check if this is a year-end Depotabschluss (December stichtag)
                                        # or early year Depotabschluss reporting on previous year
                                        if date_obj.month == 12 and date_obj.day >= 20:
                                            is_calendar_year_end = True
                                            next_calendar_year = next_cy
                                        elif date_obj.month == 1:
                                            # January Depotabschluss reporting on previous year
                                            is_calendar_year_end = True
                                            next_calendar_year = next_cy
                
                # Get or initialize counter for this fiscal year
                if fiscal_year not in fiscal_year_counters:
                    fiscal_year_counters[fiscal_year] = 0
                fiscal_year_counters[fiscal_year] += 1
                
                # Format transaction number with FY prefix
                formatted_transaction_number = f"{fiscal_year}-{fiscal_year_counters[fiscal_year]:03d}"
                
                value_date_str = self.format_date_german(value_date) if value_date else 'N/A'
                execution_date_str = self.format_date_german(execution_date) if execution_date else 'N/A'
                doc_date_str = self.format_date_german(doc_date) if doc_date else 'N/A'
                
                # Determine row class based on transaction type
                row_class = ""
                if 'Depotabschluss' in trans_type:
                    # Use different class for year-end depot statements
                    if is_year_end_depot:
                        row_class = "depot-statement-year-end"
                    else:
                        row_class = "depot-statement-row"
                    # Enrich with depot statement data
                    file_key = os.path.basename(trans.get('original_file', ''))
                    if file_key in depot_statement_lookup:
                        stmt_data = depot_statement_lookup[file_key]
                        trans['isin'] = trans.get('isin') or stmt_data.get('isin')
                        trans['shares'] = trans.get('shares') or stmt_data.get('shares')
                        # Add price per share and kurswert for depot statements
                        if not trans.get('execution_price') and stmt_data.get('price_per_share'):
                            trans['execution_price'] = stmt_data['price_per_share']
                        if not trans.get('gross_amount') and stmt_data.get('balance'):
                            trans['gross_amount'] = stmt_data['balance']
                        # Übernehme Depot-Gebühren aus Depotabschluss
                        if stmt_data.get('depot_fees') is not None and not trans.get('fees'):
                            trans['fees'] = stmt_data['depot_fees']  # Brutto
                            trans['depot_fee_net'] = stmt_data.get('depot_fee_net')
                            trans['vat_rate'] = stmt_data.get('vat_rate')
                            trans['vat_amount'] = stmt_data.get('vat_amount')
                            trans['is_depot_fee'] = True
                            trans['is_vat_free'] = False
                elif 'Verkauf' in trans_type and 'Ausführungsanzeige' not in trans_type:
                    if trans.get('profit_loss') is not None:
                        if trans['profit_loss'] > 0:
                            row_class = "profit-row"
                        elif trans['profit_loss'] < 0:
                            row_class = "loss-row"
                        else:
                            row_class = "neutral-row"
                elif trans_type == 'Kapitalmaßnahme':
                    row_class = "capital-action-row"
                elif trans_type not in ['Kauf', 'Orderabrechnung-Kauf']:
                    row_class = "misc-row"
                
                # Format display values
                isin = trans.get('isin', 'N/A')
                
                # Format shares display with balance and change columns
                shares_balance = '?'  # Default: unknown
                shares_change = EMPTY_VALUE
                
                # Get current ISIN
                current_isin = trans.get('isin', 'N/A')
                
                # First, show current balance BEFORE this transaction
                if current_isin not in ['N/A', 'None', None, '']:
                    # Valid ISIN
                    if current_isin not in isin_balances:
                        if depot_statement_seen:
                            # After depot statement: new ISINs start at 0
                            isin_balances[current_isin] = 0
                            shares_balance = '0'
                        else:
                            # Before any depot statement: unknown
                            shares_balance = '?'
                    else:
                        # ISIN is tracked - show current balance
                        shares_balance = str(isin_balances[current_isin])
                else:
                    # No valid ISIN - show dash
                    shares_balance = '-'
                
                # Now process the transaction and update balance
                if 'Depotabschluss' in trans_type:
                    # Mark that we've seen a depot statement
                    depot_statement_seen = True
                    
                    # First, set all previously tracked ISINs to 0 (they're not in depot if not mentioned)
                    for known_isin in list(isin_balances.keys()):
                        if known_isin not in ['None', 'N/A', None, '']:
                            isin_balances[known_isin] = 0
                    
                    # Then update with values from this depot statement
                    if trans.get('shares') is not None and current_isin not in ['N/A', 'None', None, '']:
                        # For depot statements, the balance shown IS the new balance
                        shares_balance = str(trans.get('shares', 0))
                        isin_balances[current_isin] = trans.get('shares', 0)
                    elif current_isin in ['None', 'N/A', None, '']:
                        # Depotabschluss without valid ISIN
                        shares_balance = EMPTY_VALUE
                    else:
                        # Specific ISIN depot statement without shares means 0 for that ISIN
                        shares_balance = '0'
                        isin_balances[current_isin] = 0
                    shares_change = EMPTY_VALUE
                    
                elif 'Kauf' in trans_type or 'Orderabrechnung-Kauf' in trans_type:
                    # Add shares on purchase (but NOT for Ausführungsanzeige)
                    if trans.get('shares'):
                        shares_change = f"+{trans['shares']}"
                        # Update balance AFTER showing it - but NOT for Ausführungsanzeige
                        if 'Ausführungsanzeige' not in trans_type:
                            if current_isin in isin_balances and isin_balances[current_isin] is not None:
                                isin_balances[current_isin] += trans['shares']
                        # If balance was unknown, it stays unknown
                    else:
                        shares_change = EMPTY_VALUE
                        
                elif 'Verkauf' in trans_type:
                    # Subtract shares on sale (but NOT for Ausführungsanzeige)
                    if trans.get('shares'):
                        shares_change = f"-{trans['shares']}"
                        # Update balance AFTER showing it - but NOT for Ausführungsanzeige
                        if 'Ausführungsanzeige' not in trans_type:
                            if current_isin in isin_balances and isin_balances[current_isin] is not None:
                                new_balance = isin_balances[current_isin] - trans['shares']
                                if new_balance < 0:
                                    print(f"  ⚠️ WARNUNG: Negative Bilanz für {current_isin} in Transaktion {transaction_number}: {isin_balances[current_isin]} - {trans['shares']} = {new_balance}")
                                    print(f"     Datum: {value_date_str}, Typ: {trans_type}, Ref: {trans.get('transaction_ref', 'N/A')}")
                                isin_balances[current_isin] = new_balance
                        # If balance was unknown, it stays unknown
                    else:
                        shares_change = EMPTY_VALUE
                        
                elif trans_type == 'Kapitalmaßnahme':
                    # Handle stock splits
                    if trans.get('split_multiplier') and trans.get('split_multiplier') > 1:
                        multiplier = trans['split_multiplier']
                        new_shares_added = trans.get('new_shares_added', 0)
                        
                        # Show the change
                        if new_shares_added > 0:
                            shares_change = f"+{new_shares_added}"  # Show actual shares added
                        else:
                            shares_change = f"×{multiplier:.0f}"  # Show multiplication factor
                        
                        # Update balance if ISIN is tracked
                        if current_isin in isin_balances and isin_balances[current_isin] is not None:
                            old_balance = isin_balances[current_isin]
                            # Calculate new balance
                            if new_shares_added > 0:
                                # Use the explicit new shares added
                                new_balance = old_balance + new_shares_added
                            else:
                                # Use multiplier
                                new_balance = int(old_balance * multiplier)
                            
                            isin_balances[current_isin] = new_balance
                            print(f"  📊 Aktiensplit {current_isin}: {old_balance} → {new_balance} (×{multiplier})")
                    else:
                        shares_change = EMPTY_VALUE
                        
                else:
                    # Other transactions don't change balance
                    shares_change = EMPTY_VALUE
                
                if trans.get('execution_price'):
                    price_str = self.format_number_german(trans['execution_price'], 2)
                else:
                    price_str = EMPTY_VALUE
                
                # Format Kurswert with sign logic
                kurswert_str = EMPTY_VALUE
                if trans.get('gross_amount'):
                    kurswert_val = trans['gross_amount']
                    if 'Verkauf' in trans_type and 'Ausführungsanzeige' not in trans_type:
                        # Sales show positive (money coming in)
                        kurswert_str = '+' + self.format_number_german(kurswert_val, 2)
                    elif 'Kauf' in trans_type:
                        # Purchases show negative (money going out)
                        kurswert_str = '-' + self.format_number_german(kurswert_val, 2)
                    else:
                        # Others without sign
                        kurswert_str = self.format_number_german(kurswert_val, 2)
                else:
                    kurswert_str = EMPTY_VALUE
                
                # Calculate fees and VAT
                gebuhren_netto_str = EMPTY_VALUE
                ust_prozent_str = EMPTY_VALUE
                ust_str = EMPTY_VALUE
                gebuhren_brutto_str = EMPTY_VALUE
                ausmachend_str = EMPTY_VALUE
                
                if trans.get('fees'):
                    # Fees are stored as gross (including VAT if applicable)
                    fees_gross = trans['fees']
                    
                    # Check if fees are VAT-free
                    if trans.get('is_vat_free', False):
                        # VAT-free: gross = net, no VAT
                        vat_rate = 0.0
                        fees_net = fees_gross
                        fees_vat = 0.0
                        ust_prozent_str = '0'
                    else:
                        # Standard fees with 19% VAT
                        vat_rate = 0.19
                        fees_net = fees_gross / (1 + vat_rate)
                        fees_vat = fees_gross - fees_net
                        ust_prozent_str = '19'
                    
                    # Store fees in transaction for later calendar year calculations
                    trans['fees_net'] = fees_net
                    trans['fees_vat'] = fees_vat
                    
                    # Fees are always costs (negative)
                    gebuhren_netto_str = '-' + self.format_number_german(fees_net, 2)
                    ust_str = '-' + self.format_number_german(fees_vat, 2) if fees_vat > 0 else '0,00'
                    gebuhren_brutto_str = '-' + self.format_number_german(fees_gross, 2)
                    
                    # Add to cumulative (fiscal year and all years)
                    cumulative_fees_net += fees_net
                    cumulative_fees_vat += fees_vat
                    all_years_fees_net += fees_net
                    all_years_fees_vat += fees_vat
                    
                    # Update calendar year cumulative values for dual tracking
                    if is_dual_tracking:
                        cumulative_cy_fees_net += fees_net
                        cumulative_cy_fees_vat += fees_vat
                else:
                    fees_gross = 0
                    fees_net = 0
                    fees_vat = 0
                    # Store zeros in transaction for later calendar year calculations
                    trans['fees_net'] = 0
                    trans['fees_vat'] = 0
                
                # Calculate Ausmachender Betrag with sign logic
                if trans.get('net_amount'):
                    ausmachend_val = trans['net_amount']
                    if 'Verkauf' in trans_type and 'Ausführungsanzeige' not in trans_type:
                        # Sales: positive if money received
                        if ausmachend_val > 0:
                            ausmachend_str = '+' + self.format_number_german(ausmachend_val, 2)
                        else:
                            ausmachend_str = self.format_number_german(ausmachend_val, 2)
                    elif 'Kauf' in trans_type:
                        # Purchases: always negative (total cost)
                        ausmachend_str = '-' + self.format_number_german(abs(ausmachend_val), 2)
                    else:
                        ausmachend_str = self.format_number_german(ausmachend_val, 2)
                elif trans.get('gross_amount'):
                    ausmachend_val = trans['gross_amount'] - fees_gross  # Subtract fees (they're positive values)
                    if 'Verkauf' in trans_type and 'Ausführungsanzeige' not in trans_type:
                        # Sales: positive if money received
                        if ausmachend_val > 0:
                            ausmachend_str = '+' + self.format_number_german(ausmachend_val, 2)
                        else:
                            ausmachend_str = self.format_number_german(ausmachend_val, 2)
                    elif 'Kauf' in trans_type:
                        # Purchases: negative (total cost = gross + fees)
                        total_cost = trans['gross_amount'] + fees_gross
                        ausmachend_str = '-' + self.format_number_german(total_cost, 2)
                    else:
                        ausmachend_str = self.format_number_german(ausmachend_val, 2)
                
                # Determine security type (stock vs non-stock)
                is_stock = False
                if isin and isin != 'N/A':
                    is_stock = self.is_stock_isin(isin)
                trans['is_stock'] = is_stock  # Store for later use in calendar year calculations
                
                # Format cumulative fee values (always negative as they are costs)
                kum_geb_str = '-' + self.format_number_german(cumulative_fees_net, 2) if cumulative_fees_net != 0 else EMPTY_VALUE
                kum_ust_str = '-' + self.format_number_german(cumulative_fees_vat, 2) if cumulative_fees_vat != 0 else EMPTY_VALUE
                
                # Format calendar year cumulative values for dual tracking
                if is_dual_tracking:
                    kum_cy_geb_str = '-' + self.format_number_german(cumulative_cy_fees_net, 2) if cumulative_cy_fees_net != 0 else EMPTY_VALUE
                    kum_cy_ust_str = '-' + self.format_number_german(cumulative_cy_fees_vat, 2) if cumulative_cy_fees_vat != 0 else EMPTY_VALUE
                
                # Initialize all G/V strings
                gv_aktien_str = EMPTY_VALUE
                kum_aktien_str = EMPTY_VALUE
                gv_non_aktien_str = EMPTY_VALUE
                kum_non_aktien_str = EMPTY_VALUE
                gv_gesamt_str = EMPTY_VALUE
                kum_gesamt_str = EMPTY_VALUE
                
                # Initialize CY cumulative strings for dual tracking
                if is_dual_tracking:
                    kum_cy_aktien_str = EMPTY_VALUE
                    kum_cy_non_aktien_str = EMPTY_VALUE
                    kum_cy_gesamt_str = EMPTY_VALUE
                
                # Calculate G/V values if profit/loss exists
                if trans.get('profit_loss') is not None:
                    profit_loss_val = trans['profit_loss']
                    
                    # Update cumulative values based on security type
                    if is_stock:
                        cumulative_stock_pnl += profit_loss_val
                        cumulative_total_pnl += profit_loss_val
                        all_years_stock_pnl += profit_loss_val
                        all_years_total_pnl += profit_loss_val
                        
                        # Update calendar year cumulative values for dual tracking
                        if is_dual_tracking:
                            cumulative_cy_stock_pnl += profit_loss_val
                            cumulative_cy_total_pnl += profit_loss_val
                        
                        # Format stock G/V
                        gv_aktien_str = self.format_number_german(profit_loss_val, 2, show_sign=True)
                        kum_aktien_str = self.format_number_german(cumulative_stock_pnl, 2, show_sign=True)
                        
                        # Format CY cumulative values for dual tracking
                        if is_dual_tracking:
                            kum_cy_aktien_str = self.format_number_german(cumulative_cy_stock_pnl, 2, show_sign=True)
                            if cumulative_cy_non_stock_pnl != 0:
                                kum_cy_non_aktien_str = self.format_number_german(cumulative_cy_non_stock_pnl, 2, show_sign=True)
                        
                        # Non-stock remains dash, but show cumulative if exists
                        if cumulative_non_stock_pnl != 0:
                            kum_non_aktien_str = self.format_number_german(cumulative_non_stock_pnl, 2, show_sign=True)
                    else:
                        cumulative_non_stock_pnl += profit_loss_val
                        cumulative_total_pnl += profit_loss_val
                        all_years_non_stock_pnl += profit_loss_val
                        all_years_total_pnl += profit_loss_val
                        
                        # Update calendar year cumulative values for dual tracking
                        if is_dual_tracking:
                            cumulative_cy_non_stock_pnl += profit_loss_val
                            cumulative_cy_total_pnl += profit_loss_val
                        
                        # Format non-stock G/V
                        gv_non_aktien_str = self.format_number_german(profit_loss_val, 2, show_sign=True)
                        kum_non_aktien_str = self.format_number_german(cumulative_non_stock_pnl, 2, show_sign=True)
                        
                        # Format CY cumulative values for dual tracking
                        if is_dual_tracking:
                            kum_cy_non_aktien_str = self.format_number_german(cumulative_cy_non_stock_pnl, 2, show_sign=True)
                            if cumulative_cy_stock_pnl != 0:
                                kum_cy_aktien_str = self.format_number_german(cumulative_cy_stock_pnl, 2, show_sign=True)
                        
                        # Stock remains dash, but show cumulative if exists
                        if cumulative_stock_pnl != 0:
                            kum_aktien_str = self.format_number_german(cumulative_stock_pnl, 2, show_sign=True)
                    
                    # Always show total G/V
                    gv_gesamt_str = self.format_number_german(profit_loss_val, 2, show_sign=True)
                    kum_gesamt_str = self.format_number_german(cumulative_total_pnl, 2, show_sign=True)
                    
                    # Format CY total cumulative for dual tracking
                    if is_dual_tracking:
                        kum_cy_gesamt_str = self.format_number_german(cumulative_cy_total_pnl, 2, show_sign=True)
                else:
                    # No G/V for this transaction, but show cumulative values if they exist
                    if cumulative_stock_pnl != 0:
                        kum_aktien_str = self.format_number_german(cumulative_stock_pnl, 2, show_sign=True)
                    if cumulative_non_stock_pnl != 0:
                        kum_non_aktien_str = self.format_number_german(cumulative_non_stock_pnl, 2, show_sign=True)
                    if cumulative_total_pnl != 0:
                        kum_gesamt_str = self.format_number_german(cumulative_total_pnl, 2, show_sign=True)
                    
                    # Format CY cumulative values if they exist
                    if is_dual_tracking:
                        if cumulative_cy_stock_pnl != 0:
                            kum_cy_aktien_str = self.format_number_german(cumulative_cy_stock_pnl, 2, show_sign=True)
                        if cumulative_cy_non_stock_pnl != 0:
                            kum_cy_non_aktien_str = self.format_number_german(cumulative_cy_non_stock_pnl, 2, show_sign=True)
                        if cumulative_cy_total_pnl != 0:
                            kum_cy_gesamt_str = self.format_number_german(cumulative_cy_total_pnl, 2, show_sign=True)
                
                # Shorten document name
                doc_name = trans.get('original_file', '')
                if len(doc_name) > 30:
                    doc_name = doc_name[:27] + '...'
                
                doc_link = f'<a href="docs/{analysis["depot_info"]["folder"]}/{trans.get("original_file", "")}">{doc_name}</a>'
                
                # Get reference number (order number or transaction ref)
                reference_str = EMPTY_VALUE
                if trans.get('order_number'):
                    reference_str = trans.get('order_number')
                elif trans.get('transaction_ref'):
                    reference_str = trans.get('transaction_ref')
                
                # Shorten type display
                if 'Orderabrechnung-' in trans_type:
                    display_type = trans_type.replace('Orderabrechnung-', '')
                elif 'Depotabschluss-' in trans_type:
                    display_type = 'Depotabschluss'
                else:
                    display_type = trans_type
                
                html += f"""
            <tr class="{row_class}">
                <td>{formatted_transaction_number}</td>
                <td>{execution_date_str}</td>
                <td>{reference_str}</td>
                <td>{value_date_str}</td>
                <td>{doc_date_str}</td>
                <td class="type-column">{display_type}</td>
                <td>{isin}</td>
                <td class="number">{shares_balance}</td>
                <td class="number">{shares_change}</td>
                <td class="number">{price_str}</td>
                <td class="number">{kurswert_str}</td>
                <td class="number fee-block-start">{gebuhren_netto_str}</td>
                <td class="number">{ust_prozent_str}</td>
                <td class="number">{ust_str}</td>
                <td class="number">{gebuhren_brutto_str}</td>
                <td class="number">{ausmachend_str}</td>"""
                
                # Add cumulative columns based on fiscal type
                if is_dual_tracking:
                    html += f"""
                <td class="number fy-column">{kum_geb_str}</td>
                <td class="number cy-column">{kum_cy_geb_str}</td>
                <td class="number fy-column">{kum_ust_str}</td>
                <td class="number fee-block-end cy-column">{kum_cy_ust_str}</td>
                <td class="number">{gv_aktien_str}</td>
                <td class="number fy-column">{kum_aktien_str}</td>
                <td class="number cy-column">{kum_cy_aktien_str}</td>
                <td class="number non-stock-block-start">{gv_non_aktien_str}</td>
                <td class="number fy-column">{kum_non_aktien_str}</td>
                <td class="number non-stock-block-end cy-column">{kum_cy_non_aktien_str}</td>"""
                else:
                    html += f"""
                <td class="number fy-column">{kum_geb_str}</td>
                <td class="number fee-block-end fy-column">{kum_ust_str}</td>
                <td class="number">{gv_aktien_str}</td>
                <td class="number fy-column">{kum_aktien_str}</td>
                <td class="number non-stock-block-start">{gv_non_aktien_str}</td>
                <td class="number non-stock-block-end fy-column">{kum_non_aktien_str}</td>"""
                
                html += f"""
                <td>{doc_link}</td>
            </tr>"""
                
                # Insert year summary rows after year-end depot statement
                if is_year_end_depot:
                    # Save the depot statement's ISIN and balance for the summary rows
                    depot_isin = isin if isin not in ['N/A', 'None', None, ''] else ''
                    depot_shares = shares_balance if shares_balance not in ['?', '-'] else ''
                    
                    # Calculate next fiscal year
                    if fiscal_year.startswith('FY'):
                        try:
                            current_year = int(fiscal_year[2:])  # Extract year from FY2021 format
                            next_fiscal_year = f'FY{current_year + 1}'
                        except (ValueError, IndexError):
                            # Fallback for unexpected format
                            next_fiscal_year = fiscal_year + '_NEXT'
                    else:
                        # Fallback for unexpected format
                        next_fiscal_year = fiscal_year + '_NEXT'
                    
                    # Get the last transaction number for this fiscal year
                    last_fy_number = fiscal_year_counters.get(fiscal_year, transaction_number)
                    
                    # Insert year-end summary row
                    html += f"""
            <tr class="year-summary-row">
                <td>{fiscal_year}-{str(last_fy_number + 1).zfill(3)}</td>
                <td colspan="5" style="text-align: center;">
                    === JAHRESSALDO {fiscal_year} ===
                </td>
                <td>{depot_isin}</td>
                <td class="number">{depot_shares}</td>
                <td></td>
                <td></td>
                <td></td>
                <td class="number fee-block-start"></td>
                <td></td>
                <td></td>
                <td></td>
                <td></td>"""
                    
                    # Add cumulative columns - only FY columns for FY transition rows
                    if is_dual_tracking:
                        html += f"""
                <td class="number bold-value">{'-' + self.format_number_german(cumulative_fees_net, 2) if cumulative_fees_net != 0 else '0,00'}</td>
                <td></td>
                <td class="number bold-value">{'-' + self.format_number_german(cumulative_fees_vat, 2) if cumulative_fees_vat != 0 else '0,00'}</td>
                <td class="number fee-block-end"></td>
                <td></td>
                <td class="number bold-value">{self.format_number_german(cumulative_stock_pnl, 2, show_sign=True) if cumulative_stock_pnl != 0 else '0,00'}</td>
                <td></td>
                <td class="non-stock-block-start"></td>
                <td class="number bold-value">{self.format_number_german(cumulative_non_stock_pnl, 2, show_sign=True) if cumulative_non_stock_pnl != 0 else '0,00'}</td>
                <td class="non-stock-block-end"></td>"""
                    else:
                        html += f"""
                <td class="number bold-value">{'-' + self.format_number_german(cumulative_fees_net, 2) if cumulative_fees_net != 0 else '0,00'}</td>
                <td class="number fee-block-end bold-value">{'-' + self.format_number_german(cumulative_fees_vat, 2) if cumulative_fees_vat != 0 else '0,00'}</td>
                <td></td>
                <td class="number bold-value">{self.format_number_german(cumulative_stock_pnl, 2, show_sign=True) if cumulative_stock_pnl != 0 else '0,00'}</td>
                <td class="non-stock-block-start"></td>
                <td class="number bold-value non-stock-block-end">{self.format_number_german(cumulative_non_stock_pnl, 2, show_sign=True) if cumulative_non_stock_pnl != 0 else '0,00'}</td>"""
                    
                    html += """
            </tr>"""
                    
                    # Insert year-start row for new fiscal year
                    # COMMENTED OUT FOR NOW - can be re-enabled later
                    # html += f"""
            # <tr class="year-start-row">
            #     <td>{next_fiscal_year}-000</td>
            #     <td colspan="5" style="text-align: center; font-style: italic;">
            #         === JAHRESANFANG {next_fiscal_year} ===
            #     </td>
            #     <td>{depot_isin}</td>
            #     <td class="number">{depot_shares}</td>
            #     <td></td>
            #     <td></td>
            #     <td></td>
            #     <td class="number fee-block-start"></td>
            #     <td></td>
            #     <td></td>
            #     <td></td>
            #     <td></td>"""
                    #
                    # # Add reset columns - only FY columns for FY transition rows
                    # if is_dual_tracking:
                    #     html += f"""
            #     <td class="number bold-value">0,00</td>
            #     <td></td>
            #     <td class="number bold-value">0,00</td>
            #     <td class="number fee-block-end"></td>
            #     <td></td>
            #     <td class="number bold-value">0,00</td>
            #     <td></td>
            #     <td class="non-stock-block-start"></td>
            #     <td class="number bold-value">0,00</td>
            #     <td class="non-stock-block-end"></td>"""
                    # else:
                    #     html += f"""
            #     <td class="number bold-value">0,00</td>
            #     <td class="number fee-block-end bold-value">0,00</td>
            #     <td></td>
            #     <td class="number bold-value">0,00</td>
            #     <td class="non-stock-block-start"></td>
            #     <td class="number bold-value non-stock-block-end">0,00</td>"""
                    #
                    # html += """
            # </tr>"""
                    
                    # Save fiscal year summary before resetting
                    fiscal_year_summaries[fiscal_year] = {
                        'stock_pnl': cumulative_stock_pnl,
                        'non_stock_pnl': cumulative_non_stock_pnl,
                        'total_pnl': cumulative_total_pnl,
                        'fees_net': cumulative_fees_net,
                        'fees_vat': cumulative_fees_vat,
                        'transactions': fiscal_year_counters.get(fiscal_year, 0),
                        'purchases_count': sum(1 for t in analysis['transactions'] 
                                             if t.get('fiscal_year') == fiscal_year 
                                             and 'Kauf' in t.get('type', '')),
                        'purchases_volume': sum((t.get('gross_amount') or 0) for t in analysis['transactions'] 
                                              if t.get('fiscal_year') == fiscal_year 
                                              and 'Kauf' in t.get('type', '')),
                        'sales_count': sum(1 for t in analysis['transactions'] 
                                         if t.get('fiscal_year') == fiscal_year 
                                         and 'Verkauf' in t.get('type', '')),
                        'sales_volume': sum((t.get('gross_amount') or 0) for t in analysis['transactions'] 
                                          if t.get('fiscal_year') == fiscal_year 
                                          and 'Verkauf' in t.get('type', ''))
                    }
                    
                    # Reset cumulative values for new fiscal year
                    cumulative_stock_pnl = 0.0
                    cumulative_non_stock_pnl = 0.0
                    cumulative_total_pnl = 0.0
                    cumulative_fees_net = 0.0
                    cumulative_fees_vat = 0.0
                
                # Insert calendar year transition rows (only for dual tracking)
                if is_calendar_year_end and is_dual_tracking:
                    calendar_year = trans.get('calendar_year', '')
                    
                    # Save calendar year summary before inserting rows
                    if calendar_year in calendar_year_summaries:
                        calendar_year_summaries[calendar_year].update({
                            'cy_stock_pnl': cumulative_cy_stock_pnl,
                            'cy_non_stock_pnl': cumulative_cy_non_stock_pnl,
                            'cy_total_pnl': cumulative_cy_total_pnl,
                            'cy_fees_net': cumulative_cy_fees_net,
                            'cy_fees_vat': cumulative_cy_fees_vat,
                            'transactions': calendar_year_counters.get(calendar_year, 0)
                        })
                    else:
                        calendar_year_summaries[calendar_year] = {
                            'cy_stock_pnl': cumulative_cy_stock_pnl,
                            'cy_non_stock_pnl': cumulative_cy_non_stock_pnl,
                            'cy_total_pnl': cumulative_cy_total_pnl,
                            'cy_fees_net': cumulative_cy_fees_net,
                            'cy_fees_vat': cumulative_cy_fees_vat,
                            'transactions': calendar_year_counters.get(calendar_year, 0),
                            'purchases_count': sum(1 for t in analysis['transactions'] 
                                                 if t.get('calendar_year') == calendar_year 
                                                 and 'Kauf' in t.get('type', '')),
                            'purchases_volume': sum((t.get('gross_amount') or 0) for t in analysis['transactions'] 
                                                  if t.get('calendar_year') == calendar_year 
                                                  and 'Kauf' in t.get('type', '')),
                            'sales_count': sum(1 for t in analysis['transactions'] 
                                             if t.get('calendar_year') == calendar_year 
                                             and 'Verkauf' in t.get('type', '')),
                            'sales_volume': sum((t.get('gross_amount') or 0) for t in analysis['transactions'] 
                                              if t.get('calendar_year') == calendar_year 
                                              and 'Verkauf' in t.get('type', ''))
                        }
                    
                    # Get transaction number for CY summary row
                    last_cy_number = calendar_year_counters.get(calendar_year, 0)
                    
                    # Insert calendar year-end summary row (blue)
                    html += f"""
            <tr class="calendar-year-summary-row">
                <td>{calendar_year}-001</td>
                <td colspan="5" style="text-align: center;">
                    === KALENDERJAHR {calendar_year.replace('CY', '')} ENDE ===
                </td>
                <td>{isin if isin not in ['N/A', 'None', None, ''] else ''}</td>
                <td class="number">{shares_balance if shares_balance not in ['?', '-'] else ''}</td>
                <td></td>
                <td></td>
                <td></td>
                <td class="number fee-block-start"></td>
                <td></td>
                <td></td>
                <td></td>
                <td></td>"""
                    
                    # Add cumulative columns - only CY columns for CY transition rows (white text on blue background)
                    html += f"""
                <td></td>
                <td class="number bold-value">{'-' + self.format_number_german(cumulative_cy_fees_net, 2) if cumulative_cy_fees_net != 0 else '0,00'}</td>
                <td></td>
                <td class="number fee-block-end bold-value">{'-' + self.format_number_german(cumulative_cy_fees_vat, 2) if cumulative_cy_fees_vat != 0 else '0,00'}</td>
                <td></td>
                <td></td>
                <td class="number bold-value">{self.format_number_german(cumulative_cy_stock_pnl, 2, show_sign=True) if cumulative_cy_stock_pnl != 0 else '0,00'}</td>
                <td class="number non-stock-block-start"></td>
                <td></td>
                <td class="number non-stock-block-end bold-value">{self.format_number_german(cumulative_cy_non_stock_pnl, 2, show_sign=True) if cumulative_cy_non_stock_pnl != 0 else '0,00'}</td>
            </tr>"""
                    
                    # Insert calendar year-start row for new year
                    # COMMENTED OUT FOR NOW - can be re-enabled later
                    # html += f"""
            # <tr class="calendar-year-start-row">
            #     <td>{next_calendar_year}-000</td>
            #     <td colspan="5" style="text-align: center; font-style: italic;">
            #         === KALENDERJAHR {next_calendar_year.replace('CY', '')} ANFANG ===
            #     </td>
            #     <td>{isin if isin not in ['N/A', 'None', None, ''] else ''}</td>
            #     <td class="number">{shares_balance if shares_balance not in ['?', '-'] else ''}</td>
            #     <td></td>
            #     <td></td>
            #     <td></td>
            #     <td class="number fee-block-start"></td>
            #     <td></td>
            #     <td></td>
            #     <td></td>
            #     <td></td>"""
                    #
                    # # Add reset cumulative columns - only CY columns for CY transition rows (white text on blue background)
                    # html += f"""
            #     <td></td>
            #     <td class="number bold-value">0,00</td>
            #     <td></td>
            #     <td class="number fee-block-end bold-value">0,00</td>
            #     <td></td>
            #     <td></td>
            #     <td class="number bold-value">0,00</td>
            #     <td class="number non-stock-block-start"></td>
            #     <td></td>
            #     <td class="number non-stock-block-end bold-value">0,00</td>
            # </tr>"""
                    
                    # Note: CY cumulative values are now reset at the beginning of the next transaction
                    # when we detect a calendar year change
                
                # Update previous calendar year for next iteration (to detect transitions)
                if is_dual_tracking:
                    previous_calendar_year = calendar_year
            
            # Save summaries for ALL calendar years (for dual tracking)
            if is_dual_tracking:
                # Collect all unique calendar years from transactions
                all_calendar_years = set()
                for t in analysis['transactions']:
                    cy = t.get('calendar_year')
                    if cy:
                        all_calendar_years.add(cy)
                
                # Create summaries for each calendar year
                for cy in all_calendar_years:
                    if cy not in calendar_year_summaries:
                        calendar_year_summaries[cy] = {}
                    
                    # Calculate values for this calendar year
                    cy_transactions = [t for t in analysis['transactions'] if t.get('calendar_year') == cy]
                    
                    # Calculate P&L for this calendar year
                    cy_stock_pnl = sum((t.get('profit_loss') or 0) for t in cy_transactions 
                                      if 'Verkauf' in t.get('type', '') and t.get('is_stock'))
                    cy_non_stock_pnl = sum((t.get('profit_loss') or 0) for t in cy_transactions 
                                          if 'Verkauf' in t.get('type', '') and not t.get('is_stock'))
                    cy_total_pnl = cy_stock_pnl + cy_non_stock_pnl
                    
                    # Calculate fees for this calendar year
                    cy_fees_net = sum((t.get('fees_net') or 0) for t in cy_transactions)
                    cy_fees_vat = sum((t.get('fees_vat') or 0) for t in cy_transactions)
                    
                    calendar_year_summaries[cy].update({
                    'cy_stock_pnl': cy_stock_pnl,
                    'cy_non_stock_pnl': cy_non_stock_pnl,
                    'cy_total_pnl': cy_total_pnl,
                    'cy_fees_net': cy_fees_net,
                    'cy_fees_vat': cy_fees_vat,
                    'transactions': calendar_year_counters.get(cy, 0),
                    'purchases_count': sum(1 for t in cy_transactions 
                                         if 'Kauf' in t.get('type', '')),
                    'purchases_volume': sum((t.get('gross_amount') or 0) for t in cy_transactions 
                                          if 'Kauf' in t.get('type', '')),
                    'sales_count': sum(1 for t in cy_transactions 
                                     if 'Verkauf' in t.get('type', '')),
                    'sales_volume': sum((t.get('gross_amount') or 0) for t in cy_transactions 
                                      if 'Verkauf' in t.get('type', ''))
                })
            
            # Save the last fiscal year's summary
            if fiscal_year:
                fiscal_year_summaries[fiscal_year] = {
                    'stock_pnl': cumulative_stock_pnl,
                    'non_stock_pnl': cumulative_non_stock_pnl,
                    'total_pnl': cumulative_total_pnl,
                    'fees_net': cumulative_fees_net,
                    'fees_vat': cumulative_fees_vat,
                    'transactions': fiscal_year_counters.get(fiscal_year, 0),
                    'purchases_count': sum(1 for t in analysis['transactions'] 
                                         if t.get('fiscal_year') == fiscal_year 
                                         and 'Kauf' in t.get('type', '')),
                    'purchases_volume': sum(t.get('gross_amount', 0) for t in analysis['transactions'] 
                                          if t.get('fiscal_year') == fiscal_year 
                                          and 'Kauf' in t.get('type', '')),
                    'sales_count': sum(1 for t in analysis['transactions'] 
                                     if t.get('fiscal_year') == fiscal_year 
                                     and 'Verkauf' in t.get('type', '')),
                    'sales_volume': sum(t.get('gross_amount', 0) for t in analysis['transactions'] 
                                      if t.get('fiscal_year') == fiscal_year 
                                      and 'Verkauf' in t.get('type', ''))
                }
            
            html += """
        </tbody>
    </table>
"""
            
            # Performance summary - now using all_years totals
            total_purchases = sum(1 for t in analysis['transactions'] if 'Kauf' in t['type'])
            total_purchases_volume = sum((t.get('gross_amount') or 0) for t in analysis['transactions'] if 'Kauf' in t['type'])
            total_sales = sum(1 for t in analysis['transactions'] if 'Verkauf' in t['type'])
            total_sales_volume = sum((t.get('gross_amount') or 0) for t in analysis['transactions'] if 'Verkauf' in t['type'])
            
            if total_sales > 0 or fiscal_year_summaries:
                # Determine header text based on fiscal year type
                year_header = "Jahr" if fiscal_type == 'calendar' else "Geschäftsjahr"
                performance_title = "Performance-Übersicht" if fiscal_type == 'calendar' else "Performance nach Geschäftsjahren"
                
                html += f"""
    <h3>{performance_title}</h3>
    <table style="width: auto;">
        <thead>
            <tr>
                <th>{year_header}</th>
                <th>Käufe<br>Anz.</th>
                <th>Käufe<br>Volumen</th>
                <th>Verkäufe<br>Anz.</th>
                <th>Verkäufe<br>Volumen</th>
                <th>G/V<br>Aktien</th>
                <th>G/V<br>Nicht-Aktien</th>
                <th>G/V<br>Gesamt</th>
                <th>Gebühren<br>(netto)</th>
                <th>USt</th>
            </tr>
        </thead>
        <tbody>"""
                
                # Add rows for each fiscal year
                if fiscal_year_summaries:
                    for fy in sorted(fiscal_year_summaries.keys()):
                        summary = fiscal_year_summaries[fy]
                        stock_pnl_class = 'positive' if summary.get('stock_pnl', 0) >= 0 else 'negative'
                        non_stock_pnl_class = 'positive' if summary.get('non_stock_pnl', 0) >= 0 else 'negative'
                        total_pnl_class = 'positive' if summary['total_pnl'] >= 0 else 'negative'
                        
                        # Format year display based on fiscal type
                        year_display = fy.replace('FY', '') if fiscal_type == 'calendar' else fy
                        
                        html += f"""
            <tr>
                <td>{year_display}</td>
                <td class="number">{summary.get('purchases_count', 0)}</td>
                <td class="number">{self.format_number_german(summary.get('purchases_volume', 0), 2)}</td>
                <td class="number">{summary.get('sales_count', 0)}</td>
                <td class="number">{self.format_number_german(summary.get('sales_volume', 0), 2)}</td>
                <td class="number {stock_pnl_class}">{self.format_number_german(summary.get('stock_pnl', 0), 2, show_sign=True)}</td>
                <td class="number {non_stock_pnl_class}">{self.format_number_german(summary.get('non_stock_pnl', 0), 2, show_sign=True)}</td>
                <td class="number {total_pnl_class}">{self.format_number_german(summary['total_pnl'], 2, show_sign=True)}</td>
                <td class="number negative">{self.format_number_german(-summary['fees_net'], 2)}</td>
                <td class="number negative">{self.format_number_german(-summary['fees_vat'], 2)}</td>
            </tr>"""
                
                # Add summary row with totals
                stock_pnl_total_class = 'positive' if all_years_stock_pnl >= 0 else 'negative'
                non_stock_pnl_total_class = 'positive' if all_years_non_stock_pnl >= 0 else 'negative'
                total_pnl_total_class = 'positive' if all_years_total_pnl >= 0 else 'negative'
                
                html += f"""
            <tr class="summary-row" style="border-top: 2px solid #333; font-weight: bold;">
                <td>Gesamt</td>
                <td class="number">{total_purchases}</td>
                <td class="number">{self.format_number_german(total_purchases_volume, 2)}</td>
                <td class="number">{total_sales}</td>
                <td class="number">{self.format_number_german(total_sales_volume, 2)}</td>
                <td class="number {stock_pnl_total_class}">{self.format_number_german(all_years_stock_pnl, 2, show_sign=True)}</td>
                <td class="number {non_stock_pnl_total_class}">{self.format_number_german(all_years_non_stock_pnl, 2, show_sign=True)}</td>
                <td class="number {total_pnl_total_class}">{self.format_number_german(all_years_total_pnl, 2, show_sign=True)}</td>
                <td class="number negative">{self.format_number_german(-all_years_fees_net, 2)}</td>
                <td class="number negative">{self.format_number_german(-all_years_fees_vat, 2)}</td>
            </tr>"""
                
                html += """
        </tbody>
    </table>
"""
            
            # Add calendar year performance table for dual tracking
            if is_dual_tracking and calendar_year_summaries:
                html += """
    <h3>Performance nach Kalenderjahren</h3>
    <table style="width: auto;">
        <thead>
            <tr>
                <th>Kalenderjahr</th>
                <th>Käufe<br>Anz.</th>
                <th>Käufe<br>Volumen</th>
                <th>Verkäufe<br>Anz.</th>
                <th>Verkäufe<br>Volumen</th>
                <th>G/V<br>Aktien</th>
                <th>G/V<br>Nicht-Aktien</th>
                <th>G/V<br>Gesamt</th>
                <th>Gebühren<br>(netto)</th>
                <th>USt</th>
            </tr>
        </thead>
        <tbody>"""
                
                # Calculate totals for calendar years
                cy_total_purchases = 0
                cy_total_purchases_volume = 0.0
                cy_total_sales = 0
                cy_total_sales_volume = 0.0
                cy_total_stock_pnl = 0.0
                cy_total_non_stock_pnl = 0.0
                cy_total_pnl = 0.0
                cy_total_fees_net = 0.0
                cy_total_fees_vat = 0.0
                
                for cy in sorted(calendar_year_summaries.keys()):
                    summary = calendar_year_summaries[cy]
                    cy_display = cy.replace('CY', '')
                    
                    # Get values with fallback
                    stock_pnl = summary.get('cy_stock_pnl', 0)
                    non_stock_pnl = summary.get('cy_non_stock_pnl', 0)
                    total_pnl = summary.get('cy_total_pnl', 0)
                    fees_net = summary.get('cy_fees_net', 0)
                    fees_vat = summary.get('cy_fees_vat', 0)
                    
                    # Update totals
                    cy_total_purchases += summary.get('purchases_count', 0)
                    cy_total_purchases_volume += summary.get('purchases_volume', 0)
                    cy_total_sales += summary.get('sales_count', 0)
                    cy_total_sales_volume += summary.get('sales_volume', 0)
                    cy_total_stock_pnl += stock_pnl
                    cy_total_non_stock_pnl += non_stock_pnl
                    cy_total_pnl += total_pnl
                    cy_total_fees_net += fees_net
                    cy_total_fees_vat += fees_vat
                    
                    # Determine CSS classes
                    stock_pnl_class = 'positive' if stock_pnl >= 0 else 'negative'
                    non_stock_pnl_class = 'positive' if non_stock_pnl >= 0 else 'negative'
                    total_pnl_class = 'positive' if total_pnl >= 0 else 'negative'
                    
                    html += f"""
            <tr>
                <td>{cy_display}</td>
                <td class="number">{summary.get('purchases_count', 0)}</td>
                <td class="number">{self.format_number_german(summary.get('purchases_volume', 0), 2)}</td>
                <td class="number">{summary.get('sales_count', 0)}</td>
                <td class="number">{self.format_number_german(summary.get('sales_volume', 0), 2)}</td>
                <td class="number {stock_pnl_class}">{self.format_number_german(stock_pnl, 2, show_sign=True)}</td>
                <td class="number {non_stock_pnl_class}">{self.format_number_german(non_stock_pnl, 2, show_sign=True)}</td>
                <td class="number {total_pnl_class}">{self.format_number_german(total_pnl, 2, show_sign=True)}</td>
                <td class="number negative">{self.format_number_german(-fees_net, 2) if fees_net > 0 else '0,00'}</td>
                <td class="number negative">{self.format_number_german(-fees_vat, 2) if fees_vat > 0 else '0,00'}</td>
            </tr>"""
                
                # Add summary row
                cy_stock_pnl_class = 'positive' if cy_total_stock_pnl >= 0 else 'negative'
                cy_non_stock_pnl_class = 'positive' if cy_total_non_stock_pnl >= 0 else 'negative'
                cy_total_pnl_class = 'positive' if cy_total_pnl >= 0 else 'negative'
                
                html += f"""
            <tr class="summary-row" style="border-top: 2px solid #333; font-weight: bold;">
                <td>Gesamt</td>
                <td class="number">{cy_total_purchases}</td>
                <td class="number">{self.format_number_german(cy_total_purchases_volume, 2)}</td>
                <td class="number">{cy_total_sales}</td>
                <td class="number">{self.format_number_german(cy_total_sales_volume, 2)}</td>
                <td class="number {cy_stock_pnl_class}">{self.format_number_german(cy_total_stock_pnl, 2, show_sign=True)}</td>
                <td class="number {cy_non_stock_pnl_class}">{self.format_number_german(cy_total_non_stock_pnl, 2, show_sign=True)}</td>
                <td class="number {cy_total_pnl_class}">{self.format_number_german(cy_total_pnl, 2, show_sign=True)}</td>
                <td class="number negative">{self.format_number_german(-cy_total_fees_net, 2) if cy_total_fees_net > 0 else '0,00'}</td>
                <td class="number negative">{self.format_number_german(-cy_total_fees_vat, 2) if cy_total_fees_vat > 0 else '0,00'}</td>
            </tr>"""
                
                html += """
        </tbody>
    </table>
"""
        
        # Add footer
        html += self._create_html_footer()
        
        return html
    
    def generate_excel(self, analysis: Dict[str, Any], output_file: str) -> None:
        """Generiert Excel-Datei mit allen Analysedaten auf separaten Tabellenblättern"""
        wb = openpyxl.Workbook()
        
        # Styles definieren
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        subheader_font = Font(bold=True)
        subheader_fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 1. Übersicht Sheet
        ws_overview = wb.active
        ws_overview.title = "Übersicht"
        
        # Kontoinfos
        depot_info = analysis.get('depot_info', {})
        company_name = analysis.get('company_name', depot_info.get('company', 'Unbekannt'))
        depot_number = analysis.get('depot_number', depot_info.get('account_number', 'N/A'))
        
        ws_overview['A1'] = f"{company_name} - Depotkonto {depot_number}"
        ws_overview['A1'].font = Font(bold=True, size=14)
        ws_overview.merge_cells('A1:D1')
        
        ws_overview['A3'] = "Kontonummer:"
        ws_overview['B3'] = depot_number
        ws_overview['A4'] = "Inhaber:"
        ws_overview['B4'] = company_name
        
        # Zusammenfassung
        ws_overview['A6'] = "Zusammenfassung"
        ws_overview['A6'].font = subheader_font
        ws_overview['A7'] = "Anzahl Transaktionen:"
        ws_overview['B7'] = len(analysis['transactions'])
        ws_overview['A8'] = "Anzahl Depotabschlüsse:"
        ws_overview['B8'] = len(analysis.get('depot_statements', []))
        
        if analysis.get('isins'):
            for idx, (isin, data) in enumerate(analysis['isins'].items(), start=10):
                ws_overview[f'A{idx}'] = f"ISIN {isin}:"
                ws_overview[f'B{idx}'] = f"{data['count']} Transaktionen"
                if data.get('total_sales_value'):
                    ws_overview[f'C{idx}'] = f"Verkaufssumme: {self.format_number_german(data['total_sales_value'], 2)} EUR"
                if data.get('total_purchase_value'):
                    ws_overview[f'D{idx}'] = f"Kaufsumme: {self.format_number_german(data['total_purchase_value'], 2)} EUR"
        
        # 2. Performance FY Sheet
        if analysis.get('fiscal_years'):
            ws_fy = wb.create_sheet("Performance FY")
            
            # Headers
            headers_fy = ["Geschäftsjahr", "Käufe", "Kaufvolumen", "Verkäufe", "Verkaufsvolumen", 
                         "G/V Aktien", "G/V Nicht-Aktien", "G/V Gesamt", "Gebühren netto", "USt"]
            for col, header in enumerate(headers_fy, 1):
                cell = ws_fy.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Data
            row = 2
            for fy_year, fy_data in sorted(analysis['fiscal_years'].items()):
                ws_fy.cell(row=row, column=1, value=fy_year)
                ws_fy.cell(row=row, column=2, value=fy_data.get('purchases', 0))
                ws_fy.cell(row=row, column=3, value=fy_data.get('purchases_volume', 0))
                ws_fy.cell(row=row, column=4, value=fy_data.get('sales', 0))
                ws_fy.cell(row=row, column=5, value=fy_data.get('sales_volume', 0))
                ws_fy.cell(row=row, column=6, value=fy_data.get('stock_pnl', 0))
                ws_fy.cell(row=row, column=7, value=fy_data.get('non_stock_pnl', 0))
                ws_fy.cell(row=row, column=8, value=fy_data.get('total_pnl', 0))
                ws_fy.cell(row=row, column=9, value=-fy_data.get('fees_net', 0) if fy_data.get('fees_net', 0) > 0 else 0)
                ws_fy.cell(row=row, column=10, value=-fy_data.get('fees_vat', 0) if fy_data.get('fees_vat', 0) > 0 else 0)
                row += 1
            
            # Spaltenbreite anpassen
            for col in range(1, 11):
                ws_fy.column_dimensions[get_column_letter(col)].width = 15
        
        # 3. Performance CY Sheet
        if analysis.get('calendar_years'):
            ws_cy = wb.create_sheet("Performance CY")
            
            # Headers
            headers_cy = ["Kalenderjahr", "Käufe", "Kaufvolumen", "Verkäufe", "Verkaufsvolumen",
                         "G/V Aktien", "G/V Nicht-Aktien", "G/V Gesamt", "Gebühren netto", "USt"]
            for col, header in enumerate(headers_cy, 1):
                cell = ws_cy.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Data
            row = 2
            for cy_year, cy_data in sorted(analysis['calendar_years'].items()):
                ws_cy.cell(row=row, column=1, value=cy_year)
                ws_cy.cell(row=row, column=2, value=cy_data.get('purchases', 0))
                ws_cy.cell(row=row, column=3, value=cy_data.get('purchases_volume', 0))
                ws_cy.cell(row=row, column=4, value=cy_data.get('sales', 0))
                ws_cy.cell(row=row, column=5, value=cy_data.get('sales_volume', 0))
                ws_cy.cell(row=row, column=6, value=cy_data.get('stock_pnl', 0))
                ws_cy.cell(row=row, column=7, value=cy_data.get('non_stock_pnl', 0))
                ws_cy.cell(row=row, column=8, value=cy_data.get('total_pnl', 0))
                ws_cy.cell(row=row, column=9, value=-cy_data.get('fees_net', 0) if cy_data.get('fees_net', 0) > 0 else 0)
                ws_cy.cell(row=row, column=10, value=-cy_data.get('fees_vat', 0) if cy_data.get('fees_vat', 0) > 0 else 0)
                row += 1
            
            # Spaltenbreite anpassen
            for col in range(1, 11):
                ws_cy.column_dimensions[get_column_letter(col)].width = 15
        
        # 4. Depotabschlüsse Sheet
        if analysis.get('depot_statements'):
            ws_depot = wb.create_sheet("Depotabschlüsse")
            
            # Headers
            headers_depot = ["Typ", "Stichtag", "Dokumentdatum", "Stück", "ISIN", "Kurs (EUR)",
                           "Depotbestand (EUR)", "Gebühren netto", "USt %", "USt (EUR)", 
                           "Gebühren brutto", "Dokument"]
            for col, header in enumerate(headers_depot, 1):
                cell = ws_depot.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Data
            row = 2
            for statement in analysis['depot_statements']:
                ws_depot.cell(row=row, column=1, value=statement.get('type_display', ''))
                ws_depot.cell(row=row, column=2, value=statement.get('balance_date', ''))
                ws_depot.cell(row=row, column=3, value=statement.get('doc_date', ''))
                ws_depot.cell(row=row, column=4, value=statement.get('shares', ''))
                ws_depot.cell(row=row, column=5, value=statement.get('isin', ''))
                ws_depot.cell(row=row, column=6, value=statement.get('price_per_share', ''))
                ws_depot.cell(row=row, column=7, value=statement.get('closing_balance', ''))
                ws_depot.cell(row=row, column=8, value=statement.get('depot_fee_net', ''))
                ws_depot.cell(row=row, column=9, value=statement.get('vat_rate', ''))
                ws_depot.cell(row=row, column=10, value=statement.get('vat_amount', ''))
                ws_depot.cell(row=row, column=11, value=statement.get('depot_fees', ''))
                ws_depot.cell(row=row, column=12, value=statement.get('file', ''))
                row += 1
            
            # Spaltenbreite anpassen
            for col in range(1, 13):
                ws_depot.column_dimensions[get_column_letter(col)].width = 15
        
        # 5. Transaktionen Sheet (Haupttabelle mit 27 Spalten)
        ws_trans = wb.create_sheet("Transaktionen")
        
        # Determine if we have dual tracking (FY != CY)
        fiscal_type = analysis.get('fiscal_year', {}).get('type', 'calendar')
        is_dual_tracking = fiscal_type == 'april_march'
        
        # Headers für 27-Spalten-Tabelle (angepasst an HTML-Tabelle)
        if is_dual_tracking:
            headers_trans = [
                "Nr", "Transaktion", "Referenz", "Wertst.", "Dokument", "Typ", "ISIN", 
                "Bestand", "Veränderung", "Kurs", "Kurswert",
                "Gebühren", "USt %", "USt", "Geb. Brutto", "Ausmach.",
                "Kum. Geb. (FY)", "Kum. Geb. (CY)", "Kum. USt (FY)", "Kum. USt (CY)",
                "G/V Akt.", "Kum. Akt. (FY)", "Kum. Akt. (CY)",
                "G/V N-Akt.", "Kum. N-Akt. (FY)", "Kum. N-Akt. (CY)", "Dokument"
            ]
        else:
            headers_trans = [
                "Nr", "Transaktion", "Referenz", "Wertst.", "Dokument", "Typ", "ISIN", 
                "Bestand", "Veränderung", "Kurs", "Kurswert",
                "Gebühren", "USt %", "USt", "Geb. Brutto", "Ausmach.",
                "Kum. Geb.", "Kum. USt",
                "G/V Akt.", "Kum. Akt.",
                "G/V N-Akt.", "Kum. N-Akt.", 
                "", "", "", "", "Dokument"  # Empty columns to maintain 27 column structure
            ]
        
        for col, header in enumerate(headers_trans, 1):
            cell = ws_trans.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Use enriched transaction data
        row = 2
        enriched_trans = analysis.get('enriched_transactions', [])
        if not enriched_trans:
            # Fallback to regular transactions if enriched not available
            enriched_trans = analysis.get('transactions', [])
        
        for trans in enriched_trans:
            # Format dates for display
            exec_date = trans.get('execution_date', trans.get('date', ''))
            if exec_date and isinstance(exec_date, str) and len(exec_date) >= 10:
                exec_date = exec_date[:10]  # Extract YYYY-MM-DD
            
            value_date = trans.get('value_date', '')
            if value_date and isinstance(value_date, str) and len(value_date) >= 10:
                value_date = value_date[:10]
            
            doc_date = trans.get('doc_date', trans.get('date', ''))
            if doc_date and isinstance(doc_date, str) and len(doc_date) >= 10:
                doc_date = doc_date[:10]
            
            # Write data based on dual tracking mode
            ws_trans.cell(row=row, column=1, value=trans.get('nr', row-1))
            ws_trans.cell(row=row, column=2, value=exec_date)
            ws_trans.cell(row=row, column=3, value=trans.get('order_number', ''))
            ws_trans.cell(row=row, column=4, value=value_date)
            ws_trans.cell(row=row, column=5, value=doc_date)
            ws_trans.cell(row=row, column=6, value=trans.get('type', ''))
            ws_trans.cell(row=row, column=7, value=trans.get('isin', ''))
            ws_trans.cell(row=row, column=8, value=trans.get('balance', ''))
            ws_trans.cell(row=row, column=9, value=trans.get('shares_change', trans.get('shares', '')))
            ws_trans.cell(row=row, column=10, value=trans.get('price', ''))
            ws_trans.cell(row=row, column=11, value=trans.get('net_amount', ''))
            ws_trans.cell(row=row, column=12, value=trans.get('fees_net', ''))
            ws_trans.cell(row=row, column=13, value=trans.get('vat_rate', ''))
            ws_trans.cell(row=row, column=14, value=trans.get('fees_vat', ''))
            ws_trans.cell(row=row, column=15, value=trans.get('fees', ''))
            ws_trans.cell(row=row, column=16, value=trans.get('gross_amount', ''))
            
            if is_dual_tracking:
                ws_trans.cell(row=row, column=17, value=trans.get('cum_fees_fy', ''))
                ws_trans.cell(row=row, column=18, value=trans.get('cum_fees_cy', ''))
                ws_trans.cell(row=row, column=19, value=trans.get('cum_vat_fy', ''))
                ws_trans.cell(row=row, column=20, value=trans.get('cum_vat_cy', ''))
                ws_trans.cell(row=row, column=21, value=trans.get('stock_pnl', ''))
                ws_trans.cell(row=row, column=22, value=trans.get('cum_stock_pnl_fy', ''))
                ws_trans.cell(row=row, column=23, value=trans.get('cum_stock_pnl_cy', ''))
                ws_trans.cell(row=row, column=24, value=trans.get('non_stock_pnl', ''))
                ws_trans.cell(row=row, column=25, value=trans.get('cum_non_stock_pnl_fy', ''))
                ws_trans.cell(row=row, column=26, value=trans.get('cum_non_stock_pnl_cy', ''))
            else:
                ws_trans.cell(row=row, column=17, value=trans.get('cum_fees_fy', ''))
                ws_trans.cell(row=row, column=18, value=trans.get('cum_vat_fy', ''))
                ws_trans.cell(row=row, column=19, value=trans.get('stock_pnl', ''))
                ws_trans.cell(row=row, column=20, value=trans.get('cum_stock_pnl_fy', ''))
                ws_trans.cell(row=row, column=21, value=trans.get('non_stock_pnl', ''))
                ws_trans.cell(row=row, column=22, value=trans.get('cum_non_stock_pnl_fy', ''))
                # Empty columns 23-26 for non-dual tracking
                ws_trans.cell(row=row, column=23, value='')
                ws_trans.cell(row=row, column=24, value='')
                ws_trans.cell(row=row, column=25, value='')
                ws_trans.cell(row=row, column=26, value='')
            
            ws_trans.cell(row=row, column=27, value=trans.get('original_file', trans.get('file', '')))
            row += 1
        
        # Spaltenbreite anpassen
        for col in range(1, 28):
            if col in [1, 3, 7]:  # Schmalere Spalten
                ws_trans.column_dimensions[get_column_letter(col)].width = 8
            elif col in [5, 6, 27]:  # Breitere Spalten
                ws_trans.column_dimensions[get_column_letter(col)].width = 20
            else:
                ws_trans.column_dimensions[get_column_letter(col)].width = 12
        
        # 6. Statistik Sheet
        ws_stats = wb.create_sheet("Statistik")
        
        # Transaction type statistics
        ws_stats['A1'] = "Transaktionsstatistik"
        ws_stats['A1'].font = Font(bold=True, size=14)
        ws_stats.merge_cells('A1:C1')
        
        row = 3
        ws_stats['A3'] = "Transaktionstyp"
        ws_stats['B3'] = "Anzahl"
        ws_stats['A3'].font = subheader_font
        ws_stats['B3'].font = subheader_font
        
        row = 4
        transaction_types = analysis.get('transaction_types', {})
        for type_name, count in transaction_types.items():
            ws_stats.cell(row=row, column=1, value=type_name)
            ws_stats.cell(row=row, column=2, value=count)
            row += 1
        
        # Spaltenbreite anpassen
        ws_stats.column_dimensions['A'].width = 30
        ws_stats.column_dimensions['B'].width = 15
        
        # Save Excel file
        wb.save(output_file)
        print(f"✓ Excel-Datei erstellt: {output_file}")
    
    def run(self):
        """Hauptfunktion - analysiert beide Depots und generiert Markdown-Dateien"""
        for depot_name in self.depots.keys():
            print(f"\nAnalysiere {depot_name} Depot...")
            analysis = self.analyze_depot(depot_name)
            
            if analysis:
                html_content = self.generate_html(analysis)
                output_file = self.depots[depot_name]['output_file']
                
                self.save_html(html_content, output_file)
                
                # Generate Excel file
                excel_file = output_file.replace('.html', '.xlsx')
                self.generate_excel(analysis, excel_file)
                
                self.print_summary(output_file, analysis)
                print(f"  - {len(analysis['transactions'])} Transaktionen analysiert")
            else:
                print(f"✗ Fehler bei der Analyse von {depot_name}")


if __name__ == "__main__":
    analyzer = DepotkontoAnalyzer()
    analyzer.run()